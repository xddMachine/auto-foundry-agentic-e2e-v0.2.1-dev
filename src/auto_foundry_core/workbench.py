"""Read-only data-room access and run-local prepared assets.

The workbench is intentionally a small boundary around :mod:`zipfile` and
streaming format readers.  It does not persist an extracted archive, build a
dataframe abstraction, or infer business meaning.  Selected members may be
materialized briefly in private temporary files for engines such as Arrow and
SQLite; the files are bounded, hash-checked, and removed before the operation
returns.  Analysts receive deterministic physical metadata, bounded rows, and
document excerpts while all durable derived state stays inside one
``RunContext``.
"""

from __future__ import annotations

import csv
from contextlib import contextmanager
from dataclasses import dataclass, field
import hashlib
import io
import json
import os
from pathlib import Path, PurePosixPath
import shutil
import sqlite3
import stat
import struct
import tempfile
from types import MappingProxyType
from typing import Any, Iterable, Iterator, Mapping, Sequence
import uuid
from urllib.parse import quote
import warnings
import zipfile

try:  # pragma: no cover - supported macOS/POSIX hosts provide fcntl
    import fcntl
except ImportError:  # pragma: no cover - defensive for non-POSIX packaging
    fcntl = None  # type: ignore[assignment]

from .contracts import DataAssetRef, PreparedAssetDescriptor
from .prepared import PreparedAssetRegistry
from .telemetry import TelemetryRecorder
from .workspace import AllowedRootError, RunContext


# JSON is parsed from a hash-checked temporary member when no explicit byte
# cap is supplied.  The old 16 MiB default was an arbitrary business limit and
# made otherwise valid local/live sources fail before an owner could inspect
# them.  Callers may still opt into a cap through ``max_json_bytes``.
DEFAULT_JSON_MAX_BYTES: int | None = None
# A caller may still provide an explicit member byte cap.  Normal Data Room
# admission does not impose an arbitrary per-member business limit; selected
# members are streamed to a secure temporary file and bounded by available
# disk/resource checks instead.
DEFAULT_MEMBER_MAX_BYTES: int | None = None
# Normal Data Room admission remains resource-safe through bounded streaming
# and available-disk checks rather than an arbitrary aggregate byte cap.  A
# caller may still provide an explicit total cap when a bounded operation is
# required.
DEFAULT_TOTAL_MAX_BYTES: int | None = None
DEFAULT_COMPRESSION_RATIO = 1000.0
DEFAULT_CATALOG_ROWS = 100_000
# XLSX XML parts can compress substantially more than ordinary text files;
# retain the compression-ratio guard, while byte/count caps remain explicit
# opt-ins rather than arbitrary default admission gates.
DEFAULT_XLSX_MEMBER_MAX_BYTES: int | None = None
DEFAULT_XLSX_TOTAL_MAX_BYTES: int | None = None
DEFAULT_XLSX_COMPRESSION_RATIO = 10_000.0
DEFAULT_XLSX_ENTRY_LIMIT: int | None = None
# Prepared candidates are bounded by the run-local disk/atomic-write and hash
# checks.  A caller may still provide ``max_bytes``/``max_output_bytes`` for a
# deliberately bounded operation.
DEFAULT_PREPARED_MAX_BYTES: int | None = None
# Version 2 records Parquet/SQLite table semantics and opaque extensionless
# members introduced by the native format catalog.  Existing version-1
# catalogs are intentionally not reused under the new identity key.
CATALOG_SCHEMA_VERSION = "2"
INSTRUMENTATION_SCHEMA_VERSION = "1"
_ZIP_EOCD_FIXED_BYTES = 22
_ZIP_EOCD_MAX_COMMENT_BYTES = 65_535
_ZIP64_EOCD_FIXED_BYTES = 56
_ZIP64_EOCD_LOCATOR_BYTES = 20
# Canonical physical metadata must not depend on a caller's sampling profile.
# Keep one bounded scan budget for every catalog identity; derived sample and
# category views use their own explicit limits later.
CANONICAL_CATALOG_ROWS = DEFAULT_CATALOG_ROWS

TABULAR_FORMATS = frozenset({"csv", "tsv", "json", "jsonl", "ndjson", "xlsx", "parquet", "sqlite"})
DOCUMENT_FORMATS = frozenset(
    {
        "txt",
        "text",
        "md",
        "markdown",
        "rst",
        "html",
        "htm",
        "xml",
        "log",
        "yaml",
        "yml",
        "toml",
        "ini",
        "cfg",
        "sql",
        "py",
        "sh",
        "pdf",
    }
)
UNSUPPORTED_FORMATS = frozenset({"xls", "feather", "avro", "doc", "docx", "ppt", "pptx"})
_RESERVED_LINEAGE_KEYS = frozenset({"archive", "members", "transformations"})


def _jsonable(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (tuple, list, set, frozenset)):
        return [_jsonable(item) for item in value]
    if isinstance(value, (Path, PurePosixPath)):
        return str(value)
    return value


def _freeze_mapping(value: Mapping[str, Any] | None) -> Mapping[str, Any]:
    return MappingProxyType({str(key): _jsonable(item) for key, item in (value or {}).items()})


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _source_stat_signature(path: Path) -> dict[str, int]:
    """Return a cheap identity check used between explicit full hashes."""

    stat_result = path.stat()
    return {
        "device": int(stat_result.st_dev),
        "inode": int(stat_result.st_ino),
        "size_bytes": int(stat_result.st_size),
        "mtime_ns": int(stat_result.st_mtime_ns),
    }


def _central_directory_fingerprint(path: Path) -> dict[str, Any]:
    """Hash ZIP central-directory bytes without enumerating ``ZipInfo`` rows.

    The stdlib's ZIP64 reader resolves the fixed ZIP64 end record immediately
    before its locator.  Mirror that layout here so the central-directory
    identity check accepts valid ZIP64 archives while still validating every
    offset before reading the claimed directory.
    """

    file_size = int(path.stat().st_size)
    if file_size < _ZIP_EOCD_FIXED_BYTES:
        raise ValueError("ZIP central directory end record is missing")
    # EOCD is at most 22 bytes plus a 65,535-byte archive comment from EOF.
    tail_size = min(file_size, _ZIP_EOCD_FIXED_BYTES + _ZIP_EOCD_MAX_COMMENT_BYTES)
    with path.open("rb") as stream:
        stream.seek(file_size - tail_size)
        tail = stream.read(tail_size)
    marker = tail.rfind(b"PK\x05\x06")
    if marker < 0 or marker + _ZIP_EOCD_FIXED_BYTES > len(tail):
        raise ValueError("ZIP central directory end record is missing")
    fields = struct.unpack_from("<4s4H2LH", tail, marker)
    _, disk_number, central_disk, disk_entries, total_entries, central_size_32, central_offset_32, comment_length = fields
    if marker + _ZIP_EOCD_FIXED_BYTES + comment_length != len(tail):
        raise ValueError("ZIP central directory comment is truncated or has trailing bytes")
    eocd_offset = file_size - tail_size + marker
    if disk_number != 0 or central_disk != 0 or disk_entries != total_entries:
        raise ValueError("multi-disk ZIP archives are not supported")

    zip64_needed = (
        disk_entries == 0xFFFF
        or total_entries == 0xFFFF
        or central_size_32 == 0xFFFFFFFF
        or central_offset_32 == 0xFFFFFFFF
    )
    locator_offset = eocd_offset - _ZIP64_EOCD_LOCATOR_BYTES
    locator: tuple[int, int, int] | None = None
    if locator_offset >= 0:
        with path.open("rb") as stream:
            stream.seek(locator_offset)
            locator_bytes = stream.read(_ZIP64_EOCD_LOCATOR_BYTES)
        if len(locator_bytes) == _ZIP64_EOCD_LOCATOR_BYTES and locator_bytes[:4] == b"PK\x06\x07":
            if not zip64_needed:
                raise ValueError("ambiguous ZIP64 end records")
            _signature, locator_disk, locator_reloff, total_disks = struct.unpack("<4sLQL", locator_bytes)
            if locator_disk != 0 or total_disks != 1:
                raise ValueError("multi-disk ZIP archives are not supported")
            locator = (locator_disk, locator_reloff, total_disks)
    if zip64_needed and locator is None:
        raise ValueError("ZIP64 end records are missing")

    if locator is not None:
        _locator_disk, locator_reloff, _total_disks = locator
        # The stdlib ZIP64 reader resolves the fixed 56-byte end record
        # immediately before the locator.  Match that mature layout and
        # reject ambiguous extensible records rather than guessing a second
        # central-directory boundary.
        zip64_offset = locator_offset - _ZIP64_EOCD_FIXED_BYTES
        if zip64_offset < 0:
            raise ValueError("ZIP64 end-of-central-directory bounds are invalid")
        with path.open("rb") as stream:
            stream.seek(zip64_offset)
            zip64_bytes = stream.read(_ZIP64_EOCD_FIXED_BYTES)
        if len(zip64_bytes) != _ZIP64_EOCD_FIXED_BYTES:
            raise ValueError("ZIP64 end-of-central-directory record is truncated")
        (
            signature,
            record_size,
            _create_version,
            _read_version,
            zip64_disk,
            zip64_central_disk,
            disk_entries,
            total_entries,
            central_size,
            central_offset,
        ) = struct.unpack("<4sQ2H2L4Q", zip64_bytes)
        if signature != b"PK\x06\x06" or record_size != 44:
            raise ValueError("ZIP64 end-of-central-directory record is malformed")
        if zip64_disk != 0 or zip64_central_disk != 0 or disk_entries != total_entries:
            raise ValueError("multi-disk ZIP archives are not supported")
        concat = zip64_offset - central_size - central_offset
        if concat < 0 or locator_reloff + concat != zip64_offset:
            raise ValueError("ZIP64 locator offset does not match the end record")
        central_end = zip64_offset
    else:
        central_size = int(central_size_32)
        central_offset = int(central_offset_32)
        central_end = eocd_offset
        concat = central_end - central_size - central_offset
        if concat < 0:
            raise ValueError("ZIP central directory offset is unsafe")

    if central_size < total_entries * 46:
        raise ValueError("ZIP central directory is too small for its entry count")
    if central_end > file_size or central_size > central_end:
        raise ValueError("ZIP central directory bounds are invalid")
    central_start = central_end - central_size
    if central_start < 0 or central_offset + concat != central_start or central_start + central_size != central_end:
        raise ValueError("ZIP central directory offset is unsafe")
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        stream.seek(central_start)
        remaining = central_size
        while remaining:
            chunk = stream.read(min(1024 * 1024, remaining))
            if not chunk:
                raise ValueError("ZIP central directory is truncated")
            digest.update(chunk)
            remaining -= len(chunk)
    return {
        "offset": int(central_offset),
        "size_bytes": int(central_size),
        "entry_count": int(total_entries),
        "sha256": digest.hexdigest(),
    }


def _atomic_write_bytes(path: Path, content: bytes) -> None:
    """Atomically publish bytes below a run-owned directory."""

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile("wb", dir=path.parent, prefix=f".{path.name}.", delete=False) as stream:
            temporary = Path(stream.name)
            stream.write(content)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
        try:
            directory = os.open(path.parent, os.O_RDONLY)
        except OSError:
            return
        try:
            try:
                os.fsync(directory)
            except OSError:
                pass
        finally:
            os.close(directory)
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def _atomic_copy_file(source: Path, destination: Path) -> None:
    """Atomically copy a temporary materialization without buffering it."""

    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.is_symlink():
        raise ValueError(f"refusing to overwrite symlink: {destination}")
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile("wb", dir=destination.parent, prefix=f".{destination.name}.", delete=False) as stream:
            temporary = Path(stream.name)
            with source.open("rb") as origin:
                shutil.copyfileobj(origin, stream, length=1024 * 1024)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, destination)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def _instrumentation_path(context: RunContext) -> Path:
    return context.resolve_run_path(Path("telemetry") / "inventory_counters.json")


@contextmanager
def _instrumentation_lock(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.with_name(f".{path.name}.lock").open("a+b") as stream:
        if fcntl is not None:
            fcntl.flock(stream.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            if fcntl is not None:
                fcntl.flock(stream.fileno(), fcntl.LOCK_UN)


def _instrumentation_warning(context: RunContext | None, reason: str) -> None:
    """Expose a passive telemetry recovery without failing the data path."""

    run_id = getattr(context, "run_id", "unknown")
    warnings.warn(
        f"inventory instrumentation recovered/ignored for run {run_id}: {reason}",
        RuntimeWarning,
        stacklevel=3,
    )


def _normalise_instrumentation_operations(value: Any) -> tuple[dict[str, dict[str, Any]], str | None]:
    """Keep only sane counter rows from a best-effort telemetry payload."""

    if not isinstance(value, Mapping):
        return {}, "operations are not an object"
    operations: dict[str, dict[str, Any]] = {}
    invalid = False
    for key, raw in value.items():
        if not isinstance(raw, Mapping):
            invalid = True
            continue
        try:
            count = int(raw.get("count", 0))
            bytes_processed = int(raw.get("bytes", 0))
        except (TypeError, ValueError, OverflowError):
            invalid = True
            continue
        if count < 0 or bytes_processed < 0:
            invalid = True
            continue
        row = dict(raw)
        row["count"] = count
        row["bytes"] = bytes_processed
        operations[str(key)] = row
    return operations, ("one or more operation counters were invalid" if invalid else None)


def _record_instrumentation(context: RunContext | None, operation: str, *, bytes_processed: int = 0) -> None:
    """Persist bounded operation counters without recording source contents."""

    if context is None:
        return
    try:
        processed = int(bytes_processed)
    except (TypeError, ValueError, OverflowError):
        _instrumentation_warning(context, "bytes_processed is invalid; counter update ignored")
        return
    if processed < 0:
        _instrumentation_warning(context, "bytes_processed is negative; counter update ignored")
        return
    try:
        path = _instrumentation_path(context)
        with _instrumentation_lock(path):
            payload: dict[str, Any] = {
                "schema_version": INSTRUMENTATION_SCHEMA_VERSION,
                "run_id": context.run_id,
                "operations": {},
            }
            diagnostics: list[str] = []
            prior_diagnostics: list[str] = []
            if path.is_file():
                try:
                    loaded = json.loads(path.read_text(encoding="utf-8"))
                except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
                    loaded = None
                    diagnostics.append(f"instrumentation is unreadable ({exc.__class__.__name__})")
                if loaded is not None:
                    if not isinstance(loaded, Mapping):
                        diagnostics.append("instrumentation payload is not an object")
                    elif loaded.get("schema_version") != INSTRUMENTATION_SCHEMA_VERSION:
                        diagnostics.append("instrumentation schema is unsupported")
                    elif loaded.get("run_id") != context.run_id:
                        diagnostics.append("instrumentation run identity does not match")
                    else:
                        operations, operation_diagnostic = _normalise_instrumentation_operations(loaded.get("operations"))
                        payload["operations"] = operations
                        if operation_diagnostic:
                            diagnostics.append(operation_diagnostic)
                        previous_diagnostics = loaded.get("diagnostics", ())
                        if isinstance(previous_diagnostics, (list, tuple)):
                            # Keep the recovery history observable without
                            # re-emitting an old warning on every healthy
                            # counter increment.
                            prior_diagnostics = [str(value) for value in previous_diagnostics[-8:]]
            if diagnostics:
                # Keep the diagnostic in the recovered payload so a caller can
                # inspect it even when warnings are filtered by the host.
                payload["diagnostics"] = list(dict.fromkeys(diagnostics[-16:]))
                _instrumentation_warning(context, "; ".join(diagnostics[:2]))
            elif prior_diagnostics:
                payload["diagnostics"] = prior_diagnostics
            operation_payload = payload["operations"].setdefault(operation, {"count": 0, "bytes": 0})
            operation_payload["count"] = int(operation_payload.get("count", 0)) + 1
            operation_payload["bytes"] = int(operation_payload.get("bytes", 0)) + processed
            try:
                _atomic_write_json(path, payload)
            except Exception as exc:  # telemetry must never block ingestion
                _instrumentation_warning(context, f"counter write failed ({exc.__class__.__name__})")
    except Exception as exc:  # lock/path failures are also observational
        _instrumentation_warning(context, f"counter update ignored ({exc.__class__.__name__})")


def _catalog_identity_key(source_hash: str, core_version: str) -> str:
    """Return the stable key for one immutable physical catalog."""

    identity = "\0".join((str(source_hash), str(core_version), CATALOG_SCHEMA_VERSION))
    return hashlib.sha256(identity.encode("utf-8")).hexdigest()


def _safe_id(value: str, label: str = "identifier") -> str:
    value = str(value).strip()
    if not value or value in {".", ".."} or Path(value).name != value or "\\" in value:
        raise ValueError(f"{label} must be a simple path component")
    return value


def _assert_no_symlink_components(path: Path, *, root: Path) -> Path:
    """Validate lexical containment and reject symlink path components."""

    try:
        relative = path.relative_to(root)
    except ValueError as exc:
        raise AllowedRootError(f"path escapes bound root: {path}") from exc
    resolved_root = root.resolve(strict=False)
    resolved_path = path.resolve(strict=False)
    if not resolved_path.is_relative_to(resolved_root):
        raise AllowedRootError(f"path escapes bound root: {path}")
    current = root
    for component in relative.parts:
        current = current / component
        if current.is_symlink():
            raise AllowedRootError(f"bound path cannot use symlink: {current}")
    return path


def _format_for_name(name: str) -> str:
    suffix = PurePosixPath(name).suffix.lower().lstrip(".")
    if suffix == "ndjson":
        return "jsonl"
    if suffix in {"db", "sqlite", "sqlite3"}:
        return "sqlite"
    if suffix:
        return suffix
    # Extensionless and otherwise unrecognised regular files remain safely
    # catalogable, but are opaque until an owner explicitly materializes them.
    return "binary"


def _member_kind(fmt: str) -> str:
    if fmt in TABULAR_FORMATS:
        return "table"
    if fmt in DOCUMENT_FORMATS or fmt == "document":
        return "document"
    # Unknown/legacy binary formats remain physically addressable but are
    # deliberately opaque: no parser or semantic fields are inferred.
    return "opaque"


def _validate_member_name(name: str) -> None:
    if not name or "\x00" in name or name.startswith("/") or name.startswith("\\") or "\\" in name:
        raise ValueError(f"unsafe ZIP member name: {name!r}")
    # PurePosixPath keeps archive names in their own namespace.  Rejecting
    # dot segments as well as parent traversal prevents ambiguous references.
    parts = PurePosixPath(name).parts
    if not parts or any(part in {"", ".", ".."} for part in parts):
        raise ValueError(f"unsafe ZIP member name: {name!r}")
    if ":" in parts[0] or PurePosixPath(name).is_absolute():
        raise ValueError(f"unsafe ZIP member name: {name!r}")


def _is_ignored_archive_metadata(name: str) -> bool:
    """Return whether *name* is inert macOS archive metadata.

    ``__MACOSX`` metadata is recognized only below that directory component;
    an ordinary member named ``__MACOSX`` remains part of the data room.  A
    ``.DS_Store`` file is ignored at any depth.  Callers must perform the
    normal member safety checks before using this predicate.
    """

    path = PurePosixPath(name)
    return path.name == ".DS_Store" or "__MACOSX" in path.parts[:-1]


def _is_symlink(info: zipfile.ZipInfo) -> bool:
    mode = (info.external_attr >> 16) & 0xFFFF
    return stat.S_IFMT(mode) == stat.S_IFLNK


def _check_limits(
    info: zipfile.ZipInfo,
    *,
    max_member_bytes: int | None,
    max_compression_ratio: float,
) -> None:
    if info.file_size < 0 or info.compress_size < 0:
        raise ValueError(f"invalid ZIP member size: {info.filename}")
    if max_member_bytes is not None and info.file_size > max_member_bytes:
        raise ValueError(
            f"ZIP member exceeds max_member_bytes: {info.filename} ({info.file_size} > {max_member_bytes})"
        )
    if info.file_size and info.compress_size == 0:
        raise ValueError(f"ZIP member has an invalid compression ratio: {info.filename}")
    if info.compress_size and info.file_size / info.compress_size > max_compression_ratio:
        raise ValueError(f"ZIP member exceeds max_compression_ratio: {info.filename}")


def _open_xlsx_zip(data: bytes) -> zipfile.ZipFile:
    try:
        return zipfile.ZipFile(io.BytesIO(data), "r")
    except zipfile.BadZipFile as exc:
        raise ValueError("invalid XLSX ZIP container") from exc


def _validate_xlsx_entry_name(info: zipfile.ZipInfo, seen_names: set[str]) -> None:
    _validate_member_name(info.filename)
    if info.filename in seen_names:
        raise ValueError(f"duplicate XLSX member name: {info.filename}")
    seen_names.add(info.filename)
    if _is_symlink(info):
        raise ValueError(f"XLSX symlink entries are not supported: {info.filename}")


def _count_xlsx_entry(info: zipfile.ZipInfo, current: int, maximum: int | None) -> int:
    count = current + 1
    if maximum is not None and count > maximum:
        raise ValueError(f"XLSX exceeds max_xlsx_entries: {count} > {maximum}")
    return count


def _validate_xlsx_file_limits(
    info: zipfile.ZipInfo,
    *,
    max_member_bytes: int | None,
    max_compression_ratio: float,
) -> None:
    if info.flag_bits & 0x1:
        raise ValueError(f"encrypted XLSX members are not supported: {info.filename}")
    _check_limits(
        info,
        max_member_bytes=max_member_bytes,
        max_compression_ratio=max_compression_ratio,
    )


def _add_xlsx_size(info: zipfile.ZipInfo, current: int, maximum: int | None) -> int:
    total = current + info.file_size
    if maximum is not None and total > maximum:
        raise ValueError(f"XLSX exceeds max_xlsx_total_bytes: {total} > {maximum}")
    return total


def _preflight_xlsx_bytes(
    data: bytes,
    *,
    max_member_bytes: int | None,
    max_total_bytes: int | None,
    max_compression_ratio: float,
    max_entries: int | None,
) -> None:
    """Validate the nested ZIP container before handing bytes to openpyxl."""

    if max_entries is not None and (isinstance(max_entries, bool) or max_entries < 0):
        raise ValueError("max_xlsx_entries cannot be negative")
    nested = _open_xlsx_zip(data)
    total_size = 0
    entry_count = 0
    seen_names: set[str] = set()
    try:
        for info in nested.infolist():
            _validate_xlsx_entry_name(info, seen_names)
            entry_count = _count_xlsx_entry(info, entry_count, max_entries)
            if info.is_dir():
                continue
            _validate_xlsx_file_limits(
                info,
                max_member_bytes=max_member_bytes,
                max_compression_ratio=max_compression_ratio,
            )
            total_size = _add_xlsx_size(info, total_size, max_total_bytes)
    finally:
        nested.close()


def _physical_type(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set, frozenset)):
        return json.dumps(_jsonable(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    if value is None:
        return ""
    return value


def _pyarrow_parquet():
    try:
        import pyarrow.parquet as parquet
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("Parquet support requires the optional 'pyarrow' dependency") from exc
    return parquet


def _quote_sqlite_identifier(identifier: str) -> str:
    """Quote one SQLite identifier without allowing SQL fragments."""

    value = str(identifier)
    if not value or "\x00" in value:
        raise ValueError("SQLite identifiers must be non-empty and NUL-free")
    return '"' + value.replace('"', '""') + '"'


def _sqlite_connect(path: Path) -> sqlite3.Connection:
    """Open a database read-only and enable SQLite's query-only guard."""

    uri = f"file:{quote(str(path), safe='/:')}?mode=ro"
    connection = sqlite3.connect(uri, uri=True)
    try:
        connection.execute("PRAGMA query_only = ON")
        connection.row_factory = sqlite3.Row
    except Exception:
        connection.close()
        raise
    return connection


def _sqlite_tables(connection: sqlite3.Connection) -> tuple[str, ...]:
    rows = connection.execute(
        "SELECT name FROM sqlite_master "
        "WHERE type = 'table' AND name NOT LIKE 'sqlite_%' "
        "ORDER BY name COLLATE BINARY"
    ).fetchall()
    return tuple(str(row[0]) for row in rows)


def _sqlite_columns(connection: sqlite3.Connection, table_name: str) -> tuple[str, ...]:
    quoted = _quote_sqlite_identifier(table_name)
    rows = connection.execute(f"PRAGMA table_info({quoted})").fetchall()
    # PRAGMA table_info returns cid order; sort explicitly so catalog shape is
    # deterministic even with unusual SQLite drivers.
    ordered = sorted(rows, key=lambda row: int(row[0]))
    return tuple(str(row[1]) for row in ordered)


@dataclass(frozen=True)
class DataRoomMember:
    """A validated physical member of the supplied archive."""

    path: str
    format: str
    kind: str
    size_bytes: int
    compressed_size_bytes: int
    content_hash: str

    @property
    def name(self) -> str:
        return self.path

    @property
    def compressed_size(self) -> int:
        """Compatibility-friendly short alias for the physical ZIP size."""

        return self.compressed_size_bytes

    @property
    def member_id(self) -> str:
        return self.path

    @property
    def sha256(self) -> str:
        return self.content_hash

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "format": self.format,
            "kind": self.kind,
            "size_bytes": self.size_bytes,
            "compressed_size_bytes": self.compressed_size_bytes,
            "content_hash": self.content_hash,
        }

    @classmethod
    def from_dict(cls, data: Mapping[str, Any]) -> "DataRoomMember":
        return cls(
            path=str(data["path"]),
            format=str(data["format"]),
            kind=str(data["kind"]),
            size_bytes=int(data["size_bytes"]),
            compressed_size_bytes=int(data["compressed_size_bytes"]),
            content_hash=str(data["content_hash"]),
        )


@dataclass(frozen=True)
class DataRoomCatalogEntry:
    """Bounded physical metadata for one member/table/sheet."""

    member: DataRoomMember
    table_name: str | None = None
    sheet_name: str | None = None
    columns: tuple[str, ...] = ()
    sample_values: Mapping[str, tuple[Any, ...]] = field(default_factory=dict)
    row_count: int | None = None
    row_count_exact: bool = False
    row_count_lower_bound: int | None = None
    sample_rows: tuple[Mapping[str, Any], ...] = ()
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        object.__setattr__(self, "columns", tuple(str(value) for value in self.columns))
        object.__setattr__(
            self,
            "sample_values",
            MappingProxyType(
                {
                    str(key): tuple(_jsonable(value) for value in values)
                    for key, values in self.sample_values.items()
                }
            ),
        )
        object.__setattr__(self, "sample_rows", tuple(_freeze_mapping(row) for row in self.sample_rows))
        object.__setattr__(self, "metadata", _freeze_mapping(self.metadata))

    @property
    def path(self) -> str:
        return self.member.path

    @property
    def member_path(self) -> str:
        return self.member.path

    @property
    def source_id(self) -> str:
        """Stable logical source identity for analyst selection and search.

        SQLite databases expand into one catalog entry per user table, so the
        table name is part of the identity.  XLSX keeps its historical sheet
        spelling for callers that already persist ``path::sheet=...`` IDs.
        """

        if self.member.format == "sqlite" and self.table_name is not None:
            return f"{self.path}::table={self.table_name}"
        if self.sheet_name is not None:
            return f"{self.path}::sheet={self.sheet_name}"
        return self.path

    @property
    def catalog_id(self) -> str:
        return self.source_id

    @property
    def entry_id(self) -> str:
        return self.source_id

    @property
    def id(self) -> str:
        return self.source_id

    @property
    def table(self) -> str | None:
        return self.table_name

    @property
    def sheet(self) -> str | None:
        return self.sheet_name

    @property
    def format(self) -> str:
        return self.member.format

    @property
    def kind(self) -> str:
        return self.member.kind

    @property
    def content_hash(self) -> str:
        return self.member.content_hash

    @property
    def size_bytes(self) -> int:
        return self.member.size_bytes

    @property
    def compressed_size_bytes(self) -> int:
        return self.member.compressed_size_bytes

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_id": self.source_id,
            "member": self.member.to_dict(),
            "table_name": self.table_name,
            "sheet_name": self.sheet_name,
            "columns": list(self.columns),
            "sample_values": {key: list(values) for key, values in self.sample_values.items()},
            "row_count": self.row_count,
            "row_count_exact": self.row_count_exact,
            "row_count_lower_bound": self.row_count_lower_bound,
            "sample_rows": [_jsonable(row) for row in self.sample_rows],
            "metadata": _jsonable(self.metadata),
        }

    @classmethod
    def from_dict(cls, data: Mapping[str, Any]) -> "DataRoomCatalogEntry":
        entry = cls(
            member=DataRoomMember.from_dict(data["member"]),
            table_name=data.get("table_name"),
            sheet_name=data.get("sheet_name"),
            columns=tuple(data.get("columns", ())),
            sample_values={key: tuple(values) for key, values in dict(data.get("sample_values", {})).items()},
            row_count=data.get("row_count"),
            row_count_exact=bool(data.get("row_count_exact", False)),
            row_count_lower_bound=data.get("row_count_lower_bound"),
            sample_rows=tuple(data.get("sample_rows", ())),
            metadata=data.get("metadata", {}),
        )
        persisted_source_id = data.get("source_id")
        if persisted_source_id is not None and str(persisted_source_id) != entry.source_id:
            raise ValueError("catalog entry source_id does not match its physical/table identity")
        return entry


@dataclass(frozen=True)
class CatalogCounts:
    """Typed counts for the physical archive and expanded catalog.

    ``archive_members`` counts physical ZIP members.  ``catalog_entries``
    counts expanded member/table/sheet entries.  ``table_members`` counts
    distinct physical tabular members, while ``sheet_entries`` counts the
    expanded XLSX worksheet entries.  Keeping these values in a named
    contract prevents callers from accidentally comparing an archive count
    with an expanded catalog count by position.
    """

    archive_members: int
    catalog_entries: int
    table_members: int
    sheet_entries: int

    def __post_init__(self) -> None:
        for field_name in ("archive_members", "catalog_entries", "table_members", "sheet_entries"):
            value = getattr(self, field_name)
            if isinstance(value, bool) or not isinstance(value, int) or value < 0:
                raise TypeError(f"{field_name} must be a non-negative integer")

    def to_dict(self) -> dict[str, int]:
        return {
            "archive_members": self.archive_members,
            "catalog_entries": self.catalog_entries,
            "table_members": self.table_members,
            "sheet_entries": self.sheet_entries,
        }


@dataclass(frozen=True)
class PreparedAsset:
    """A loadable prepared asset and its integrity descriptor."""

    descriptor: PreparedAssetDescriptor
    rows: tuple[Mapping[str, Any], ...]

    def __post_init__(self) -> None:
        object.__setattr__(self, "rows", tuple(_freeze_mapping(row) for row in self.rows))

    def __iter__(self) -> Iterator[Mapping[str, Any]]:
        return iter(self.rows)

    def __len__(self) -> int:
        return len(self.rows)

    def __getitem__(self, index: int) -> Mapping[str, Any]:
        return self.rows[index]

    @property
    def prepared_asset_id(self) -> str:
        return self.descriptor.prepared_asset_id


@dataclass(frozen=True)
class _ArchiveLimits:
    max_member_bytes: int | None = DEFAULT_MEMBER_MAX_BYTES
    max_total_bytes: int | None = DEFAULT_TOTAL_MAX_BYTES
    max_compression_ratio: float = DEFAULT_COMPRESSION_RATIO
    max_json_bytes: int | None = DEFAULT_JSON_MAX_BYTES
    max_catalog_rows: int = DEFAULT_CATALOG_ROWS
    max_xlsx_member_bytes: int | None = DEFAULT_XLSX_MEMBER_MAX_BYTES
    max_xlsx_total_bytes: int | None = DEFAULT_XLSX_TOTAL_MAX_BYTES
    max_xlsx_compression_ratio: float = DEFAULT_XLSX_COMPRESSION_RATIO
    max_xlsx_entries: int | None = DEFAULT_XLSX_ENTRY_LIMIT


def _archive_limits(
    *,
    max_member_bytes: int | None,
    max_total_bytes: int | None,
    max_compression_ratio: float,
    max_json_bytes: int | None,
    max_catalog_rows: int,
    max_xlsx_member_bytes: int | None,
    max_xlsx_total_bytes: int | None,
    max_xlsx_compression_ratio: float,
    max_xlsx_entries: int | None,
) -> _ArchiveLimits:
    nonnegative = (
        max_member_bytes,
        max_total_bytes,
        max_json_bytes,
        max_catalog_rows,
        max_xlsx_member_bytes,
        max_xlsx_total_bytes,
        max_xlsx_entries,
    )
    if any(value is not None and value < 0 for value in nonnegative):
        raise ValueError("workbench bounds cannot be negative")
    if max_compression_ratio <= 0 or max_xlsx_compression_ratio <= 0:
        raise ValueError("compression ratios must be positive")
    return _ArchiveLimits(
        max_member_bytes=max_member_bytes,
        max_total_bytes=max_total_bytes,
        max_compression_ratio=max_compression_ratio,
        max_json_bytes=max_json_bytes,
        max_catalog_rows=max_catalog_rows,
        max_xlsx_member_bytes=max_xlsx_member_bytes,
        max_xlsx_total_bytes=max_xlsx_total_bytes,
        max_xlsx_compression_ratio=max_xlsx_compression_ratio,
        max_xlsx_entries=max_xlsx_entries,
    )


def _inspect_archive_member(
    source: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    limits: _ArchiveLimits,
    seen_names: set[str],
) -> DataRoomMember | None:
    _validate_member_name(info.filename)
    if info.filename in seen_names:
        raise ValueError(f"duplicate ZIP member name: {info.filename}")
    seen_names.add(info.filename)
    if _is_symlink(info):
        raise ValueError(f"ZIP symlink entries are not supported: {info.filename}")
    if info.is_dir():
        return None
    if info.flag_bits & 0x1:
        raise ValueError(f"encrypted ZIP members are not supported: {info.filename}")
    _check_limits(
        info,
        max_member_bytes=limits.max_member_bytes,
        max_compression_ratio=limits.max_compression_ratio,
    )
    if _is_ignored_archive_metadata(info.filename):
        return None
    fmt = _format_for_name(info.filename)
    digest = hashlib.sha256()
    xlsx_bytes = bytearray() if fmt == "xlsx" else None
    with source.open(info, "r") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
            if xlsx_bytes is not None:
                xlsx_bytes.extend(chunk)
    if xlsx_bytes is not None:
        _preflight_xlsx_bytes(
            bytes(xlsx_bytes),
            max_member_bytes=limits.max_xlsx_member_bytes,
            max_total_bytes=limits.max_xlsx_total_bytes,
            max_compression_ratio=limits.max_xlsx_compression_ratio,
            max_entries=limits.max_xlsx_entries,
        )
    return DataRoomMember(
        path=info.filename,
        format=fmt,
        kind=_member_kind(fmt),
        size_bytes=info.file_size,
        compressed_size_bytes=info.compress_size,
        content_hash=digest.hexdigest(),
    )


def _inventory_archive(
    archive_path: Path,
    *,
    limits: _ArchiveLimits,
    context: RunContext | None = None,
    record_instrumentation: bool = True,
) -> tuple[DataRoomMember, ...]:
    members: list[DataRoomMember] = []
    seen_names: set[str] = set()
    total_size = 0
    try:
        with zipfile.ZipFile(archive_path, "r") as source:
            for info in source.infolist():
                member = _inspect_archive_member(source, info, limits=limits, seen_names=seen_names)
                if member is None:
                    # Ignored metadata is still part of the archive's bounded
                    # physical footprint, while directory accounting remains
                    # unchanged from the pre-filter inventory.
                    if not info.is_dir() and _is_ignored_archive_metadata(info.filename):
                        total_size += info.file_size
                        if limits.max_total_bytes is not None and total_size > limits.max_total_bytes:
                            raise ValueError(f"ZIP exceeds max_total_bytes: {total_size} > {limits.max_total_bytes}")
                    continue
                total_size += member.size_bytes
                if limits.max_total_bytes is not None and total_size > limits.max_total_bytes:
                    raise ValueError(f"ZIP exceeds max_total_bytes: {total_size} > {limits.max_total_bytes}")
                members.append(member)
                if record_instrumentation:
                    _record_instrumentation(context, "member_content_hash", bytes_processed=member.size_bytes)
    except zipfile.BadZipFile as exc:
        raise ValueError(f"invalid ZIP archive: {archive_path}") from exc
    return tuple(sorted(members, key=lambda item: item.path))


class DataRoom:
    """Read-only access to one validated ZIP archive."""

    def __init__(
        self,
        context: RunContext,
        archive_path: Path,
        archive_ref: DataAssetRef,
        members: tuple[DataRoomMember, ...],
        *,
        limits: _ArchiveLimits,
        telemetry: TelemetryRecorder | None = None,
        bound_inventory: bool = False,
    ) -> None:
        self.context = context
        self.archive_path = archive_path
        self.archive_ref = archive_ref
        self._members = members
        self.limits = limits
        self.telemetry = telemetry
        self._bound_inventory = bool(bound_inventory)
        self._source_stat = _source_stat_signature(archive_path)
        # A re-bound analysis context may deliberately keep using the exact
        # catalog created by its previous implementation.  The workbench
        # installs this private override after opening the bound archive; it
        # is never inferred from the current core version.
        self._bound_catalog_entries: tuple[DataRoomCatalogEntry, ...] | None = None
        self.catalog_schema_version = CATALOG_SCHEMA_VERSION
        self.catalog_key = _catalog_identity_key(archive_ref.content_hash or "", context.core_version)
        self.catalog_root = context.resolve_run_path(Path("data_room") / "catalogs")
        self.catalog_path = context.resolve_run_path(self.catalog_root / f"{self.catalog_key}.json")
        self.catalog_lock_path = context.resolve_run_path(self.catalog_root / ".catalog.lock")

    @classmethod
    def open(
        cls,
        context: RunContext,
        archive: str | Path | DataAssetRef,
        *,
        telemetry: TelemetryRecorder | None = None,
        max_member_bytes: int | None = DEFAULT_MEMBER_MAX_BYTES,
        max_total_bytes: int | None = DEFAULT_TOTAL_MAX_BYTES,
        max_compression_ratio: float = DEFAULT_COMPRESSION_RATIO,
        max_json_bytes: int | None = DEFAULT_JSON_MAX_BYTES,
        max_catalog_rows: int = DEFAULT_CATALOG_ROWS,
        max_xlsx_member_bytes: int | None = DEFAULT_XLSX_MEMBER_MAX_BYTES,
        max_xlsx_total_bytes: int | None = DEFAULT_XLSX_TOTAL_MAX_BYTES,
        max_xlsx_compression_ratio: float = DEFAULT_XLSX_COMPRESSION_RATIO,
        max_xlsx_entries: int | None = DEFAULT_XLSX_ENTRY_LIMIT,
        bound_members: Sequence[DataRoomMember] | None = None,
        bound_archive_hash: str | None = None,
        bound_source_stat: Mapping[str, Any] | None = None,
        bound_central_directory_fingerprint: Mapping[str, Any] | None = None,
        bound_identity_only: bool = False,
        record_instrumentation: bool = True,
    ) -> "DataRoom":
        limits = _archive_limits(
            max_member_bytes=max_member_bytes,
            max_total_bytes=max_total_bytes,
            max_compression_ratio=max_compression_ratio,
            max_json_bytes=max_json_bytes,
            max_catalog_rows=max_catalog_rows,
            max_xlsx_member_bytes=max_xlsx_member_bytes,
            max_xlsx_total_bytes=max_xlsx_total_bytes,
            max_xlsx_compression_ratio=max_xlsx_compression_ratio,
            max_xlsx_entries=max_xlsx_entries,
        )
        archive_value = archive
        supplied_hash: str | None = None
        if isinstance(archive_value, DataAssetRef):
            supplied_hash = archive_value.content_hash
            archive_value = archive_value.uri
        archive_path = context.resolve_input(archive_value)
        if not archive_path.exists() or not archive_path.is_file():
            raise FileNotFoundError(archive_path)
        current_stat = _source_stat_signature(archive_path)
        expected_stat = dict(bound_source_stat or {})
        if expected_stat and current_stat != {str(key): int(value) for key, value in expected_stat.items()}:
            raise ValueError(f"source stat changed after binding: {archive_path}")
        expected_central = dict(bound_central_directory_fingerprint or {})
        if bound_identity_only and bound_members is None:
            raise ValueError("identity-only source binding requires bound physical inventory")
        if bound_identity_only:
            if not expected_central:
                raise ValueError("bound source identity must contain a central-directory fingerprint")
            # A context transition deliberately reuses the immutable physical
            # inventory and central-directory fingerprint already bound in its
            # manifest.  Only the cheap source stat is checked here; no ZIP
            # central-directory or member read/counter is allowed.
            central_fingerprint = expected_central
        else:
            central_fingerprint = _central_directory_fingerprint(archive_path)
            if bound_members is not None and not expected_central:
                raise ValueError("bound source identity must contain a central-directory fingerprint")
            if expected_central and central_fingerprint != expected_central:
                raise ValueError(f"archive central directory changed after binding: {archive_path}")
            if record_instrumentation:
                _record_instrumentation(
                    context,
                    "central_directory_fingerprint",
                    bytes_processed=int(central_fingerprint["size_bytes"]),
                )
        if bound_members is None:
            archive_hash = _sha256_file(archive_path)
            if record_instrumentation:
                _record_instrumentation(context, "archive_full_hash", bytes_processed=archive_path.stat().st_size)
            if supplied_hash and supplied_hash != archive_hash:
                raise ValueError(f"source changed after registration: {archive_path}")
            members_tuple = _inventory_archive(
                archive_path,
                limits=limits,
                context=context,
                record_instrumentation=record_instrumentation,
            )
            bound_inventory = False
        else:
            archive_hash = str(bound_archive_hash or supplied_hash or "")
            if not archive_hash:
                raise ValueError("bound source identity must contain an archive hash")
            if supplied_hash and supplied_hash != archive_hash:
                raise ValueError(f"bound source hash does not match: {archive_path}")
            members_tuple = cls._validate_bound_inventory(tuple(bound_members), limits=limits)
            bound_inventory = True
        room = cls(
            context,
            archive_path,
            DataAssetRef(
                uri=str(archive_path),
                format="zip",
                content_hash=archive_hash,
                size_bytes=archive_path.stat().st_size,
                metadata={
                    "read_only": True,
                    "source_stat": current_stat,
                    "central_directory_fingerprint": central_fingerprint,
                },
            ),
            members_tuple,
            limits=limits,
            telemetry=telemetry,
            bound_inventory=bound_inventory,
        )
        if not bound_identity_only:
            room._emit(
                "data_room_archive_read",
                bytes_processed=archive_path.stat().st_size,
                facts={"archive_hash": archive_hash, "member_count": len(members_tuple)},
            )
        return room

    @staticmethod
    def _validate_bound_inventory(
        members: tuple[DataRoomMember, ...],
        *,
        limits: _ArchiveLimits,
    ) -> tuple[DataRoomMember, ...]:
        """Validate persisted member metadata without opening the ZIP."""

        expected = {member.path: member for member in members}
        if len(expected) != len(members):
            raise ValueError("bound physical inventory contains duplicate members")
        total_size = 0
        for member in members:
            _validate_member_name(member.path)
            if member.kind not in {"table", "document", "opaque"}:
                raise ValueError(f"bound member kind is invalid: {member.path}")
            if member.size_bytes < 0 or member.compressed_size_bytes < 0:
                raise ValueError(f"bound member sizes are invalid: {member.path}")
            if not member.content_hash or len(member.content_hash) != 64:
                raise ValueError(f"bound member hash is invalid: {member.path}")
            if limits.max_member_bytes is not None and member.size_bytes > limits.max_member_bytes:
                raise ValueError(
                    f"ZIP member exceeds max_member_bytes: {member.path} "
                    f"({member.size_bytes} > {limits.max_member_bytes})"
                )
            total_size += member.size_bytes
            if limits.max_total_bytes is not None and total_size > limits.max_total_bytes:
                raise ValueError(f"ZIP exceeds max_total_bytes: {total_size} > {limits.max_total_bytes}")
        return tuple(sorted(members, key=lambda item: item.path))

    def members(self) -> tuple[DataRoomMember, ...]:
        self._check_archive_stat()
        return self._members

    def _emit(self, event_type: str, *, bytes_processed: int | None = None, facts: Mapping[str, Any] | None = None) -> None:
        if self.telemetry is None:
            return
        try:
            self.telemetry.record(
                event_type,
                bytes_processed=bytes_processed,
                facts=dict(facts or {}),
            )
        except Exception:
            # Telemetry is observational only.
            pass

    def _check_archive_unchanged(self, *, record_instrumentation: bool = True) -> None:
        current = _sha256_file(self.archive_path)
        if record_instrumentation:
            _record_instrumentation(self.context, "archive_full_hash", bytes_processed=self.archive_path.stat().st_size)
        if current != self.archive_ref.content_hash:
            raise ValueError(f"archive changed after registration: {self.archive_path}")

    def _check_archive_stat(self) -> None:
        expected = self.archive_ref.metadata.get("source_stat", self._source_stat)
        if _source_stat_signature(self.archive_path) != dict(expected):
            raise ValueError(f"archive changed after binding (source stat): {self.archive_path}")

    def verify_source_full(self, *, record_instrumentation: bool = True) -> str:
        """Explicitly re-hash the bound archive at a final/freeze boundary.

        Read-only projections can request the same strict byte validation
        without persisting observational counters.  The default remains
        instrumented so ingestion/freeze callers retain their existing
        telemetry semantics.
        """

        self._check_archive_stat()
        if record_instrumentation:
            _record_instrumentation(
                self.context,
                "verify_source_full",
                bytes_processed=self.archive_path.stat().st_size,
            )
        self._check_archive_unchanged(record_instrumentation=record_instrumentation)
        return self.archive_ref.content_hash or ""

    @property
    def source_stat_signature(self) -> Mapping[str, int]:
        return MappingProxyType(dict(self._source_stat))

    @property
    def central_directory_fingerprint(self) -> Mapping[str, Any]:
        return MappingProxyType(dict(self.archive_ref.metadata.get("central_directory_fingerprint", {})))

    @property
    def instrumentation_path(self) -> Path:
        return _instrumentation_path(self.context)

    @property
    def instrumentation_counters(self) -> Mapping[str, Any]:
        path = self.instrumentation_path
        if not path.is_file():
            return MappingProxyType({})
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            _instrumentation_warning(self.context, f"counter read ignored ({exc.__class__.__name__})")
            return MappingProxyType({})
        if not isinstance(payload, Mapping):
            _instrumentation_warning(self.context, "counter payload is not an object")
            return MappingProxyType({})
        if payload.get("schema_version") != INSTRUMENTATION_SCHEMA_VERSION:
            _instrumentation_warning(self.context, "counter schema is unsupported")
            return MappingProxyType({})
        if payload.get("run_id") != self.context.run_id:
            _instrumentation_warning(self.context, "counter run identity does not match")
            return MappingProxyType({})
        operations, diagnostic = _normalise_instrumentation_operations(payload.get("operations"))
        if diagnostic:
            _instrumentation_warning(self.context, diagnostic)
        return MappingProxyType(_jsonable(operations))

    def _resolve_member(self, member: DataRoomMember | DataRoomCatalogEntry | str | Path) -> DataRoomMember:
        expected_hash: str | None = None
        if isinstance(member, DataRoomCatalogEntry):
            expected_hash = member.member.content_hash
            member = member.member
        path = member.path if isinstance(member, DataRoomMember) else str(member)
        for candidate in self._members:
            if candidate.path == path:
                if expected_hash is not None and candidate.content_hash != expected_hash:
                    raise ValueError(f"catalog member hash changed: {path}")
                return candidate
        raise KeyError(f"unknown data-room member: {path}")

    def _zip_info(self, member: DataRoomMember) -> zipfile.ZipInfo:
        with zipfile.ZipFile(self.archive_path, "r") as source:
            try:
                info = source.getinfo(member.path)
            except KeyError as exc:
                raise KeyError(f"unknown data-room member: {member.path}") from exc
        _validate_member_name(info.filename)
        if _is_symlink(info) or info.flag_bits & 0x1:
            raise ValueError(f"unsafe ZIP member: {member.path}")
        _check_limits(
            info,
            max_member_bytes=self.limits.max_member_bytes,
            max_compression_ratio=self.limits.max_compression_ratio,
        )
        if info.file_size != member.size_bytes or info.compress_size != member.compressed_size_bytes:
            raise ValueError(f"archive member metadata changed: {member.path}")
        return info

    @contextmanager
    def _materialized_member(
        self,
        member: DataRoomMember,
        *,
        max_bytes: int | None = None,
        count_selected: bool = True,
    ) -> Iterator[Path]:
        """Stream one selected ZIP member into a private temporary file.

        The default path has no arbitrary business-size ceiling.  A caller
        supplied ``max_bytes`` remains an explicit opt-in cap.  Before any
        decompression, check actual free space on the temporary filesystem;
        then stream/hash/CRC-check in bounded chunks and remove the file on
        every exit path.
        """

        self._check_archive_stat()
        info = self._zip_info(member)
        cap = self.limits.max_member_bytes if max_bytes is None else int(max_bytes)
        if cap is not None and cap < 0:
            raise ValueError("max_bytes cannot be negative")
        if cap is not None and info.file_size > cap:
            raise ValueError(f"member exceeds max_bytes: {member.path} ({info.file_size} > {cap})")
        with tempfile.TemporaryDirectory(prefix="auto-foundry-member-") as directory:
            try:
                free_bytes = int(shutil.disk_usage(directory).free)
            except OSError as exc:
                raise ValueError("unable to determine temporary disk capacity") from exc
            if free_bytes < info.file_size:
                raise ValueError(
                    f"insufficient temporary disk space for selected member: "
                    f"{info.file_size} > {free_bytes}"
                )
            file_descriptor, temporary_name = tempfile.mkstemp(
                prefix="member-",
                suffix=PurePosixPath(member.path).suffix,
                dir=directory,
            )
            temporary_path = Path(temporary_name)
            try:
                os.fchmod(file_descriptor, 0o600)
                digest = hashlib.sha256()
                total = 0
                with os.fdopen(file_descriptor, "wb") as destination:
                    file_descriptor = -1
                    with zipfile.ZipFile(self.archive_path, "r") as source, source.open(info, "r") as stream:
                        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                            total += len(chunk)
                            if cap is not None and total > cap:
                                raise ValueError(f"member exceeds max_bytes: {member.path}")
                            digest.update(chunk)
                            destination.write(chunk)
                    destination.flush()
                    os.fsync(destination.fileno())
                if digest.hexdigest() != member.content_hash:
                    raise ValueError(f"archive member content changed: {member.path}")
                _record_instrumentation(self.context, "member_content_hash", bytes_processed=total)
                if count_selected:
                    _record_instrumentation(self.context, "selected_member_read", bytes_processed=total)
                self._emit(
                    "data_room_member_materialized",
                    bytes_processed=total,
                    facts={"member_path": member.path, "format": member.format},
                )
                yield temporary_path
            finally:
                if file_descriptor >= 0:
                    try:
                        os.close(file_descriptor)
                    except OSError:
                        pass
                temporary_path.unlink(missing_ok=True)

    def _read_member_bytes(
        self,
        member: DataRoomMember,
        *,
        max_bytes: int | None = None,
        allow_truncate: bool = False,
        count_selected: bool = True,
    ) -> bytes:
        self._check_archive_stat()
        info = self._zip_info(member)
        cap = self.limits.max_member_bytes if max_bytes is None else max_bytes
        if cap is not None and cap < 0:
            raise ValueError("max_bytes cannot be negative")
        if cap is not None and info.file_size > cap and not allow_truncate:
            raise ValueError(f"member exceeds max_bytes: {member.path} ({info.file_size} > {cap})")
        digest = hashlib.sha256()
        chunks: list[bytes] = []
        total = 0
        with zipfile.ZipFile(self.archive_path, "r") as source, source.open(info, "r") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
                previous_total = total
                total += len(chunk)
                if cap is None:
                    chunks.append(chunk)
                elif previous_total < cap:
                    chunks.append(chunk[: cap - previous_total])
        data = b"".join(chunks)
        if cap is not None and total > cap and not allow_truncate:
            raise ValueError(f"member exceeds max_bytes: {member.path}")
        if digest.hexdigest() != member.content_hash:
            raise ValueError(f"archive member content changed: {member.path}")
        _record_instrumentation(self.context, "member_content_hash", bytes_processed=total)
        if count_selected:
            _record_instrumentation(self.context, "selected_member_read", bytes_processed=total)
        self._emit(
            "data_room_member_read",
            bytes_processed=total,
            facts={"member_path": member.path, "format": member.format},
        )
        return data

    def _iter_csv_rows(
        self,
        member: DataRoomMember,
        *,
        delimiter: str,
        limit: int | None,
        offset: int,
    ) -> list[dict[str, Any]]:
        output: list[dict[str, Any]] = []
        with self._materialized_member(member) as path:
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream, delimiter=delimiter)
                for index, row in enumerate(reader):
                    if index < offset:
                        continue
                    output.append(dict(row))
                    if limit is not None and len(output) >= limit:
                        break
        self._emit(
            "data_room_member_read",
            bytes_processed=member.size_bytes,
            facts={"member_path": member.path, "format": member.format, "rows": len(output)},
        )
        return output

    def _iter_json_rows(
        self,
        member: DataRoomMember,
        *,
        limit: int | None,
        offset: int,
        max_json_bytes: int | None = None,
        count_selected: bool = True,
    ) -> list[dict[str, Any]]:
        if member.format == "json":
            json_cap = self.limits.max_json_bytes if max_json_bytes is None else max_json_bytes
            if json_cap is None:
                # Keep the no-cap path streaming at the archive boundary.  The
                # standard JSON decoder still materializes its parsed value,
                # but no arbitrary 16 MiB byte buffer is imposed before the
                # owner can inspect a valid source.
                with self._materialized_member(member, count_selected=count_selected) as path:
                    with path.open("r", encoding="utf-8-sig") as stream:
                        value = json.load(stream)
            else:
                data = self._read_member_bytes(
                    member,
                    max_bytes=json_cap,
                    count_selected=count_selected,
                )
                value = json.loads(data.decode("utf-8-sig"))
            if isinstance(value, Mapping):
                records = value.get("records", value.get("data"))
                value = records if isinstance(records, list) else [value]
            elif not isinstance(value, list):
                value = [value]
            rows = [item if isinstance(item, Mapping) else {"value": item} for item in value]
            selected = rows[offset:] if limit is None else rows[offset : offset + limit]
            return [dict(row) for row in selected]

        output: list[dict[str, Any]] = []
        row_index = 0
        with self._materialized_member(member, count_selected=count_selected) as path:
            with path.open("r", encoding="utf-8-sig") as stream:
                for line_number, line in enumerate(stream, 1):
                    if not line.strip():
                        continue
                    value = json.loads(line)
                    row = dict(value) if isinstance(value, Mapping) else {"value": value, "_line": line_number}
                    if row_index < offset:
                        row_index += 1
                        continue
                    output.append(row)
                    row_index += 1
                    if limit is not None and len(output) >= limit:
                        break
        self._emit(
            "data_room_member_read",
            bytes_processed=member.size_bytes,
            facts={"member_path": member.path, "format": member.format, "rows": len(output)},
        )
        return output

    def _xlsx_rows(
        self,
        member: DataRoomMember,
        *,
        sheet: str | None,
        limit: int | None,
        offset: int,
    ) -> list[dict[str, Any]]:
        data = self._read_member_bytes(member)
        _preflight_xlsx_bytes(
            data,
            max_member_bytes=self.limits.max_xlsx_member_bytes,
            max_total_bytes=self.limits.max_xlsx_total_bytes,
            max_compression_ratio=self.limits.max_xlsx_compression_ratio,
            max_entries=self.limits.max_xlsx_entries,
        )
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise RuntimeError("XLSX support requires the optional 'openpyxl' dependency") from exc
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            sheet_name = sheet or workbook.sheetnames[0]
            if sheet_name not in workbook.sheetnames:
                raise KeyError(f"unknown worksheet: {sheet_name}")
            values = workbook[sheet_name].iter_rows(values_only=True)
            try:
                headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(values))]
            except StopIteration:
                return []
            output: list[dict[str, Any]] = []
            for index, row in enumerate(values):
                if index < offset:
                    continue
                output.append({headers[column]: row[column] if column < len(row) else None for column in range(len(headers))})
                if limit is not None and len(output) >= limit:
                    break
            self._emit(
                "data_room_member_read",
                bytes_processed=len(data),
                facts={"member_path": member.path, "format": member.format, "sheet": sheet_name, "rows": len(output)},
            )
            return output
        finally:
            workbook.close()

    def _parquet_rows(
        self,
        member: DataRoomMember,
        *,
        limit: int | None,
        offset: int,
    ) -> list[dict[str, Any]]:
        """Read Parquet in bounded Arrow record batches from a temp file."""

        output: list[dict[str, Any]] = []
        skipped = 0
        with self._materialized_member(member) as path:
            try:
                parquet = _pyarrow_parquet().ParquetFile(path)
                for batch in parquet.iter_batches(batch_size=1024):
                    for value in batch.to_pylist():
                        if skipped < offset:
                            skipped += 1
                            continue
                        output.append(dict(value))
                        if limit is not None and len(output) >= limit:
                            break
                    if limit is not None and len(output) >= limit:
                        break
            except (OSError, ValueError, RuntimeError) as exc:
                if isinstance(exc, RuntimeError) and "pyarrow" in str(exc).lower():
                    raise
                raise ValueError(f"invalid Parquet member: {member.path}") from exc
        self._emit(
            "data_room_member_read",
            facts={"member_path": member.path, "format": member.format, "rows": len(output)},
        )
        return output

    def _sqlite_rows(
        self,
        member: DataRoomMember,
        *,
        table_name: str | None,
        limit: int | None,
        offset: int,
    ) -> list[dict[str, Any]]:
        """Read one SQLite user table using a read-only bounded cursor."""

        output: list[dict[str, Any]] = []
        skipped = 0
        with self._materialized_member(member) as path:
            try:
                connection = _sqlite_connect(path)
            except sqlite3.DatabaseError as exc:
                raise ValueError(f"invalid SQLite member: {member.path}") from exc
            try:
                tables = _sqlite_tables(connection)
                selected = table_name
                if selected is None:
                    if not tables:
                        return []
                    selected = tables[0]
                if selected not in tables:
                    raise KeyError(f"unknown SQLite table: {selected}")
                cursor = connection.execute(f"SELECT * FROM {_quote_sqlite_identifier(selected)}")
                while True:
                    batch = cursor.fetchmany(1024)
                    if not batch:
                        break
                    for row in batch:
                        if skipped < offset:
                            skipped += 1
                            continue
                        output.append({str(key): row[key] for key in row.keys()})
                        if limit is not None and len(output) >= limit:
                            break
                    if limit is not None and len(output) >= limit:
                        break
            except KeyError:
                raise
            except sqlite3.DatabaseError as exc:
                raise ValueError(f"invalid SQLite table: {member.path}") from exc
            finally:
                connection.close()
        self._emit(
            "data_room_member_read",
            facts={"member_path": member.path, "format": member.format, "table": table_name, "rows": len(output)},
        )
        return output

    def read_rows(
        self,
        member: DataRoomMember | DataRoomCatalogEntry | str | Path,
        *,
        sheet: str | None = None,
        table: str | None = None,
        limit: int | None = 1000,
        offset: int = 0,
        max_json_bytes: int | None = None,
    ) -> list[dict[str, Any]]:
        if limit is not None and limit < 0:
            raise ValueError("limit cannot be negative")
        if offset < 0:
            raise ValueError("offset cannot be negative")
        if limit == 0:
            return []
        resolved = self._resolve_member(member)
        selected_table: str | None = table
        if isinstance(member, DataRoomCatalogEntry):
            if sheet is None:
                sheet = member.sheet_name
            if selected_table is None:
                selected_table = member.table_name if member.format == "sqlite" else None
        if sheet is not None and table is not None:
            raise ValueError("sheet and table cannot both be supplied")
        if resolved.kind == "opaque":
            raise ValueError(f"opaque data-room member requires explicit materialization: {resolved.path}")
        if resolved.format == "csv":
            return self._iter_csv_rows(resolved, delimiter=",", limit=limit, offset=offset)
        if resolved.format == "tsv":
            return self._iter_csv_rows(resolved, delimiter="\t", limit=limit, offset=offset)
        if resolved.format in {"json", "jsonl"}:
            if max_json_bytes is not None:
                if max_json_bytes < 0:
                    raise ValueError("max_json_bytes cannot be negative")
            return self._iter_json_rows(
                resolved,
                limit=limit,
                offset=offset,
                max_json_bytes=max_json_bytes,
            )
        if resolved.format == "xlsx":
            return self._xlsx_rows(resolved, sheet=sheet, limit=limit, offset=offset)
        if resolved.format == "parquet":
            if sheet is not None or table is not None:
                raise ValueError("table/sheet selectors are not valid for Parquet members")
            return self._parquet_rows(resolved, limit=limit, offset=offset)
        if resolved.format == "sqlite":
            if sheet is not None:
                selected_table = sheet
            return self._sqlite_rows(resolved, table_name=selected_table, limit=limit, offset=offset)
        if sheet is not None:
            raise ValueError("sheet is only valid for XLSX members")
        raise ValueError(f"member is a document, not a table: {resolved.path}")

    def document_excerpt(
        self,
        member: DataRoomMember | DataRoomCatalogEntry | str | Path,
        *,
        max_bytes: int = 65536,
    ) -> str:
        if max_bytes < 0:
            raise ValueError("max_bytes cannot be negative")
        resolved = self._resolve_member(member)
        if resolved.kind == "opaque":
            raise ValueError(f"opaque data-room member requires explicit materialization: {resolved.path}")
        if resolved.kind != "document":
            raise ValueError(f"member is tabular, not a document: {resolved.path}")
        if resolved.format == "pdf":
            if max_bytes == 0:
                return ""
            # Reuse the bounded, isolated PDF normalizer instead of making a
            # valid PDF a hard workbench failure.  The member is fully
            # hash/CRC-checked by ``_read_member_bytes`` before pypdf sees it;
            # extraction failures remain a visible empty/limited result in the
            # document catalog rather than an exception at this boundary.
            from .document_ingestion import normalize_document_bytes

            data = self._read_member_bytes(resolved, max_bytes=None)
            normalized = normalize_document_bytes(
                data,
                document_ref=resolved.path,
                source_path=resolved.path,
                format="pdf",
                max_document_bytes=None,
                max_member_bytes=None,
                max_total_bytes=None,
                max_entries=None,
                max_excerpt_bytes=max_bytes,
                max_pdf_output_bytes=max_bytes,
            )
            text = "\n".join(section.text for section in normalized.sections)
            # The normalizer applies a UTF-8 byte cap per section.  Joining
            # pages adds separators, so enforce the public excerpt contract
            # once more without splitting a multi-byte character.
            return text.encode("utf-8")[:max_bytes].decode("utf-8", errors="ignore")
        data = self._read_member_bytes(resolved, max_bytes=max_bytes, allow_truncate=True)
        return data.decode("utf-8-sig", errors="replace")

    def materialize_opaque(
        self,
        member: DataRoomMember | DataRoomCatalogEntry | str | Path,
        destination: str | Path,
        *,
        max_bytes: int | None = None,
    ) -> Path:
        """Copy one opaque member to an explicit run-local regular file."""

        resolved = self._resolve_member(member)
        if resolved.kind != "opaque":
            raise ValueError(f"opaque materialization requires an opaque member: {resolved.path}")
        cap = self.limits.max_member_bytes if max_bytes is None else int(max_bytes)
        if cap is not None and cap < 0:
            raise ValueError("max_bytes cannot be negative")
        if cap is not None and resolved.size_bytes > cap:
            raise ValueError(f"opaque member exceeds max_bytes: {resolved.path}")
        # Check the lexical destination before resolving it.  Resolving first
        # would silently collapse a pre-existing symlink alias that still
        # lands inside ``run_root`` and would make the write path ambiguous.
        raw_destination = Path(destination).expanduser()
        lexical_target = raw_destination if raw_destination.is_absolute() else self.context.run_root / raw_destination
        _assert_no_symlink_components(lexical_target, root=self.context.run_root)
        target = self.context.resolve_run_path(lexical_target)
        if target.exists() and (target.is_symlink() or not target.is_file()):
            raise ValueError(f"opaque materialization destination must be a regular file: {target}")
        with self._materialized_member(resolved, max_bytes=cap) as temporary:
            _atomic_copy_file(temporary, target)
        if _sha256_file(target) != resolved.content_hash:
            target.unlink(missing_ok=True)
            raise ValueError(f"opaque materialization hash mismatch: {resolved.path}")
        return target

    def _catalog_for_member(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        if member.kind == "opaque":
            return [
                DataRoomCatalogEntry(
                    member=member,
                    row_count=None,
                    row_count_exact=False,
                    row_count_lower_bound=None,
                    metadata={"extraction": "opaque", "requires_custom_code": True},
                )
            ]
        if member.kind == "document":
            return self._catalog_document(member)
        if member.format == "xlsx":
            return self._catalog_xlsx(member)
        if member.format == "parquet":
            return self._catalog_parquet(member)
        if member.format == "sqlite":
            return self._catalog_sqlite(member)
        if member.format in {"json", "jsonl"}:
            return self._catalog_json(member)
        return self._catalog_delimited(member)

    def _catalog_document(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        if member.format == "pdf":
            return [
                DataRoomCatalogEntry(
                    member=member,
                    row_count=None,
                    row_count_exact=False,
                    row_count_lower_bound=None,
                    metadata={"extraction": "opaque", "requires_custom_code": True},
                )
            ]
        return [
            DataRoomCatalogEntry(
                member=member,
                row_count=None,
                row_count_exact=False,
                row_count_lower_bound=None,
            )
        ]

    def _catalog_xlsx(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        data = self._read_member_bytes(member, count_selected=False)
        _preflight_xlsx_bytes(
            data,
            max_member_bytes=self.limits.max_xlsx_member_bytes,
            max_total_bytes=self.limits.max_xlsx_total_bytes,
            max_compression_ratio=self.limits.max_xlsx_compression_ratio,
            max_entries=self.limits.max_xlsx_entries,
        )
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("XLSX support requires the optional 'openpyxl' dependency") from exc
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        entries: list[DataRoomCatalogEntry] = []
        try:
            for sheet_name in workbook.sheetnames:
                entries.append(
                    self._catalog_xlsx_sheet(
                        member,
                        workbook[sheet_name],
                        sheet_name=sheet_name,
                    )
                )
        finally:
            workbook.close()
        self._emit("data_room_member_read", bytes_processed=len(data), facts={"member_path": member.path, "format": member.format, "sheets": len(entries)})
        return entries

    def _catalog_xlsx_sheet(
        self,
        member: DataRoomMember,
        worksheet: Any,
        *,
        sheet_name: str,
    ) -> DataRoomCatalogEntry:
        values = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value) if value is not None else f"column_{index + 1}" for index, value in enumerate(next(values))]
        except StopIteration:
            headers = []
        count = 0
        exhausted = True
        for row in values:
            count += 1
            if count >= CANONICAL_CATALOG_ROWS:
                try:
                    next(values)
                except StopIteration:
                    pass
                else:
                    exhausted = False
                break
        return self._entry_from_rows(member, headers, sheet_name, count, exhausted)

    def _catalog_parquet(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        """Catalog Parquet from Arrow metadata without materializing rows."""

        with self._materialized_member(member, count_selected=False) as path:
            try:
                parquet = _pyarrow_parquet().ParquetFile(path)
                schema = parquet.schema_arrow
                columns = tuple(str(name) for name in schema.names)
                metadata = parquet.metadata
                row_count = int(metadata.num_rows) if metadata is not None else None
                row_groups = int(metadata.num_row_groups) if metadata is not None else None
            except RuntimeError:
                raise
            except Exception as exc:
                raise ValueError(f"invalid Parquet member: {member.path}") from exc
        return [
            DataRoomCatalogEntry(
                member=member,
                table_name=PurePosixPath(member.path).stem,
                columns=columns,
                row_count=row_count,
                row_count_exact=row_count is not None,
                row_count_lower_bound=row_count,
                metadata={
                    "extraction": "pyarrow",
                    "row_groups": row_groups,
                },
            )
        ]

    def _catalog_sqlite(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        """Catalog each SQLite user table from schema metadata only.

        Exact ``COUNT(*)`` is intentionally deferred to analytical profiling:
        a catalog pass must remain bounded even when a table contains billions
        of rows.  Consumers receive an explicit unknown row count and can use
        :meth:`read_rows`/the profiling layer when an exact or lower-bound
        observation is actually needed.
        """

        entries: list[DataRoomCatalogEntry] = []
        with self._materialized_member(member, count_selected=False) as path:
            try:
                connection = _sqlite_connect(path)
            except sqlite3.DatabaseError as exc:
                raise ValueError(f"invalid SQLite member: {member.path}") from exc
            try:
                table_names = _sqlite_tables(connection)
                for table_name in table_names:
                    columns = _sqlite_columns(connection, table_name)
                    entries.append(
                        DataRoomCatalogEntry(
                            member=member,
                            table_name=table_name,
                            columns=columns,
                            row_count=None,
                            row_count_exact=False,
                            row_count_lower_bound=None,
                            metadata={
                                "extraction": "sqlite",
                                "read_only": True,
                                "query_only": True,
                                "row_count_kind": "unknown",
                            },
                        )
                    )
            except sqlite3.DatabaseError as exc:
                raise ValueError(f"invalid SQLite table metadata: {member.path}") from exc
            finally:
                connection.close()
        if entries:
            return entries
        # An empty but valid database remains catalogable and non-blocking.  It
        # has no logical table rows, so expose one physical entry without
        # pretending that a user table exists.
        return [
            DataRoomCatalogEntry(
                member=member,
                table_name=None,
                columns=(),
                row_count=0,
                row_count_exact=True,
                row_count_lower_bound=0,
                metadata={"extraction": "sqlite", "read_only": True, "query_only": True, "user_tables": 0},
            )
        ]

    def _catalog_json(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        rows = self._iter_json_rows(member, limit=CANONICAL_CATALOG_ROWS + 1, offset=0, count_selected=False)
        headers: list[str] = []
        for row in rows:
            for key in row:
                if str(key) not in headers:
                    headers.append(str(key))
        exhausted = len(rows) <= CANONICAL_CATALOG_ROWS
        count = min(len(rows), CANONICAL_CATALOG_ROWS)
        return [self._entry_from_rows(member, headers, None, count, exhausted)]

    def _catalog_delimited(self, member: DataRoomMember) -> list[DataRoomCatalogEntry]:
        delimiter = "\t" if member.format == "tsv" else ","
        count = 0
        exhausted = True
        with self._materialized_member(member, count_selected=False) as path:
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream, delimiter=delimiter)
                headers = [str(value) for value in (reader.fieldnames or ())]
                for row in reader:
                    count += 1
                    if count >= CANONICAL_CATALOG_ROWS:
                        try:
                            next(reader)
                        except StopIteration:
                            pass
                        else:
                            exhausted = False
                        break
        self._emit("data_room_member_read", bytes_processed=member.size_bytes, facts={"member_path": member.path, "format": member.format, "rows": count})
        return [self._entry_from_rows(member, headers, None, count, exhausted)]

    @staticmethod
    def _entry_from_rows(
        member: DataRoomMember,
        headers: Sequence[str],
        sheet_name: str | None,
        count: int,
        exact: bool,
    ) -> DataRoomCatalogEntry:
        columns: list[str] = [str(value) for value in headers]
        return DataRoomCatalogEntry(
            member=member,
            table_name=sheet_name or PurePosixPath(member.path).stem,
            sheet_name=sheet_name,
            columns=tuple(columns),
            row_count=count if exact else None,
            row_count_exact=exact,
            row_count_lower_bound=count,
        )

    def _load_canonical_catalog(self) -> tuple[DataRoomCatalogEntry, ...]:
        """Load and validate the one immutable catalog for this source identity."""

        if not self.catalog_path.is_file():
            raise FileNotFoundError(self.catalog_path)
        try:
            payload = json.loads(self.catalog_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValueError("canonical source catalog is unreadable") from exc
        if not isinstance(payload, Mapping):
            raise ValueError("canonical source catalog must be an object")
        expected = {
            "catalog_schema_version": CATALOG_SCHEMA_VERSION,
            "catalog_key": self.catalog_key,
            "source_hash": self.archive_ref.content_hash,
            "core_version": self.context.core_version,
        }
        if any(payload.get(key) != value for key, value in expected.items()):
            raise ValueError("canonical source catalog identity does not match current run")
        raw_entries = payload.get("entries")
        if not isinstance(raw_entries, list):
            raise ValueError("canonical source catalog entries must be a list")
        entries = tuple(DataRoomCatalogEntry.from_dict(value) for value in raw_entries)
        for entry in entries:
            # Validate every persisted physical member against the immutable
            # archive inventory before exposing catalog metadata to callers.
            self._resolve_member(entry.member)
        # Canonical entries are physical metadata only.  A non-empty sample or
        # category field indicates a caller-specific view was persisted and is
        # rejected rather than silently reused as canonical state.
        if any(entry.sample_values or entry.sample_rows for entry in entries):
            raise ValueError("canonical source catalog contains derived sample data")
        raw_counts = payload.get("counts")
        if not isinstance(raw_counts, Mapping):
            raise ValueError("canonical source catalog counts must be an object")
        try:
            persisted_counts = CatalogCounts(
                archive_members=raw_counts["archive_members"],
                catalog_entries=raw_counts["catalog_entries"],
                table_members=raw_counts["table_members"],
                sheet_entries=raw_counts["sheet_entries"],
            )
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError("canonical source catalog counts are invalid") from exc
        if persisted_counts != self.catalog_counts(entries):
            raise ValueError("canonical source catalog counts do not match entries")
        return entries

    def build_catalog(self) -> tuple[DataRoomCatalogEntry, ...]:
        """Build or load one immutable physical catalog for this source hash.

        Sampling and category limits are intentionally absent.  They belong to
        the derived ``sample``/``categories`` APIs and never alter canonical
        catalog identity or trigger a full archive rescan.
        """

        self._check_archive_stat()
        if self._bound_catalog_entries is not None:
            return self._bound_catalog_entries
        if self.catalog_path.is_file():
            entries = self._load_canonical_catalog()
            _record_instrumentation(self.context, "catalog_loaded")
            _record_instrumentation(self.context, "catalog_reused")
            self._emit(
                "data_room_catalog_reused",
                facts={
                    "entry_count": len(entries),
                    "catalog_path": str(self.catalog_path),
                    "catalog_key": self.catalog_key,
                    "catalog_schema_version": CATALOG_SCHEMA_VERSION,
                },
            )
            return entries
        with _catalog_lock(self.catalog_lock_path):
            self._check_archive_stat()
            # A concurrent opener may have won publication while this caller
            # waited for the lock.  Recheck before touching any member.
            if self.catalog_path.is_file():
                entries = self._load_canonical_catalog()
                _record_instrumentation(self.context, "catalog_loaded")
                _record_instrumentation(self.context, "catalog_reused")
                self._emit(
                    "data_room_catalog_reused",
                    facts={
                        "entry_count": len(entries),
                        "catalog_path": str(self.catalog_path),
                        "catalog_key": self.catalog_key,
                        "catalog_schema_version": CATALOG_SCHEMA_VERSION,
                    },
                )
                return entries
            entries: list[DataRoomCatalogEntry] = []
            for member in self._members:
                entries.extend(self._catalog_for_member(member))
            # A source mutation during the bounded scan must fail closed and
            # must never publish a partial catalog under the old identity.
            self._check_archive_stat()
            result = tuple(sorted(entries, key=lambda item: (item.path, item.sheet_name or "")))
            counts = self.catalog_counts(result)
            payload = {
                "catalog_schema_version": CATALOG_SCHEMA_VERSION,
                "catalog_key": self.catalog_key,
                "source_hash": self.archive_ref.content_hash,
                "core_version": self.context.core_version,
                "archive": self.archive_ref.to_dict(),
                "counts": counts.to_dict(),
                "entries": [entry.to_dict() for entry in result],
            }
            _atomic_write_json(self.catalog_path, payload)
            _record_instrumentation(self.context, "catalog_created")
            self._emit(
                "data_room_catalog_created",
                facts={
                    "entry_count": len(result),
                    "catalog_path": str(self.catalog_path),
                    "catalog_key": self.catalog_key,
                    "catalog_schema_version": CATALOG_SCHEMA_VERSION,
                    **counts.to_dict(),
                },
            )
            return result

    def catalog_counts(self, catalog: Iterable[DataRoomCatalogEntry] | None = None) -> CatalogCounts:
        """Return distinct physical-member and expanded-entry counts."""

        entries = tuple(catalog) if catalog is not None else self.build_catalog()
        table_paths = {entry.member.path for entry in entries if entry.member.kind == "table"}
        sheet_entries = sum(1 for entry in entries if entry.sheet_name is not None)
        return CatalogCounts(
            archive_members=len(self._members),
            catalog_entries=len(entries),
            table_members=len(table_paths),
            sheet_entries=sheet_entries,
        )

    def sample(
        self,
        entry: DataRoomCatalogEntry | DataRoomMember | str | Path,
        *,
        limit: int = 20,
    ) -> tuple[Mapping[str, Any], ...]:
        """Read a bounded sample from one explicitly selected member/view."""

        if limit < 0:
            raise ValueError("limit cannot be negative")
        if limit == 0:
            return ()
        if isinstance(entry, DataRoomCatalogEntry):
            selected = entry
            member = entry.member
            sheet = entry.sheet_name
        else:
            selected = None
            member = self._resolve_member(entry)
            sheet = None
        if member.kind == "document":
            excerpt_limit = max(limit, 1) * 4096
            if self.limits.max_member_bytes is not None:
                excerpt_limit = min(self.limits.max_member_bytes, excerpt_limit)
            excerpt = self.document_excerpt(member, max_bytes=excerpt_limit)
            return tuple({"text": line} for line in excerpt.splitlines()[:limit])
        rows = self.read_rows(member if selected is None else selected, sheet=sheet, limit=limit)
        return tuple(_freeze_mapping(row) for row in rows)

    def categories(
        self,
        entry: DataRoomCatalogEntry | DataRoomMember | str | Path,
        column: str,
        *,
        limit: int = 20,
    ) -> tuple[Any, ...]:
        """Return bounded distinct values from one selected member/table/sheet."""

        if limit < 0:
            raise ValueError("limit cannot be negative")
        if limit == 0:
            return ()
        if isinstance(entry, DataRoomCatalogEntry):
            selected = entry
            member = entry.member
            sheet = entry.sheet_name
            if selected.columns and str(column) not in selected.columns:
                raise KeyError(f"unknown catalog column: {column}")
        else:
            selected = None
            member = self._resolve_member(entry)
            sheet = None
        if member.kind != "table":
            raise ValueError(f"categories require a tabular member: {member.path}")
        values: list[Any] = []
        rows = self.read_rows(member if selected is None else selected, sheet=sheet, limit=self.limits.max_catalog_rows)
        for row in rows:
            value = row.get(column)
            if value in (None, "") or value in values:
                continue
            values.append(_jsonable(value))
            if len(values) >= limit:
                break
        return tuple(values)

    def search(
        self,
        query: str,
        *,
        catalog: Iterable[DataRoomCatalogEntry] | None = None,
        limit: int = 20,
    ) -> tuple[DataRoomCatalogEntry, ...]:
        if limit < 0:
            raise ValueError("limit cannot be negative")
        entries = tuple(catalog) if catalog is not None else self.build_catalog()
        tokens = tuple(token for token in str(query).strip().lower().split() if token)
        if not tokens:
            return tuple(sorted(entries, key=lambda item: (item.path, item.sheet_name or "")))[:limit]
        ranked: list[tuple[int, DataRoomCatalogEntry]] = []
        for entry in entries:
            values = [
                entry.path,
                entry.source_id,
                entry.format,
                entry.kind,
                entry.table_name or "",
                entry.sheet_name or "",
                *entry.columns,
            ]
            haystack = " ".join(values).lower()
            score = sum(1 for token in tokens if token in haystack)
            if score:
                ranked.append((score, entry))
        ranked.sort(key=lambda pair: (-pair[0], pair[1].path, pair[1].sheet_name or ""))
        return tuple(entry for _, entry in ranked[:limit])


def _atomic_write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    encoded = json.dumps(_jsonable(payload), ensure_ascii=False, sort_keys=True, indent=2) + "\n"
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", delete=False) as stream:
            temporary = Path(stream.name)
            stream.write(encoded)
            stream.flush()
            os.fsync(stream.fileno())
        temporary.replace(path)
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


@contextmanager
def _catalog_lock(path: Path):
    """Serialize canonical catalog creation across concurrent callers."""

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a+b") as stream:
        if fcntl is not None:
            fcntl.flock(stream.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            if fcntl is not None:
                fcntl.flock(stream.fileno(), fcntl.LOCK_UN)


@contextmanager
def _candidate_lock(path: Path):
    """Serialize candidate publication within one item workspace."""

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a+b") as stream:
        if fcntl is not None:
            fcntl.flock(stream.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            if fcntl is not None:
                fcntl.flock(stream.fileno(), fcntl.LOCK_UN)


class DataRoomWorkbench:
    """Analyst-facing workbench for one archive and one run context."""

    def __init__(
        self,
        context: RunContext,
        archive: str | Path | DataAssetRef,
        *,
        runtime: Any = None,
        telemetry: TelemetryRecorder | None = None,
        _bound_members: Sequence[DataRoomMember] | None = None,
        _bound_archive_hash: str | None = None,
        _bound_source_stat: Mapping[str, Any] | None = None,
        _bound_central_directory_fingerprint: Mapping[str, Any] | None = None,
        _bound_catalog_entries: Sequence[DataRoomCatalogEntry] | None = None,
        _bound_catalog_path: str | Path | None = None,
        _bound_catalog_key: str | None = None,
        **limits: Any,
    ) -> None:
        if runtime is not None and getattr(runtime, "context", context) is not context:
            raise ValueError("runtime must use the same RunContext as the workbench")
        self.context = context
        self.runtime = runtime
        self.telemetry = telemetry or getattr(runtime, "telemetry", None) or TelemetryRecorder(context=context)
        self._room = DataRoom.open(
            context,
            archive,
            telemetry=self.telemetry,
            bound_members=_bound_members,
            bound_archive_hash=_bound_archive_hash,
            bound_source_stat=_bound_source_stat,
            bound_central_directory_fingerprint=_bound_central_directory_fingerprint,
            bound_identity_only=_bound_catalog_entries is not None,
            **limits,
        )
        if _bound_catalog_entries is not None:
            self._room._bound_catalog_entries = tuple(_bound_catalog_entries)
            if _bound_catalog_path is not None:
                self._room.catalog_path = Path(_bound_catalog_path)
            if _bound_catalog_key is not None:
                self._room.catalog_key = str(_bound_catalog_key)
        self.prepared_registry = PreparedAssetRegistry(context, telemetry=self.telemetry)
        self._prepared: dict[str, PreparedAssetDescriptor] = {}

    @property
    def data_room(self) -> DataRoom:
        return self._room

    def catalog(self) -> tuple[DataRoomCatalogEntry, ...]:
        return self._room.build_catalog()

    def _resolve_source_member(self, value: DataRoomMember | DataRoomCatalogEntry | str | Path) -> DataRoomMember:
        return self._room._resolve_member(value)

    def _normalize_source_refs(
        self,
        source_refs: Iterable[DataAssetRef | str],
    ) -> tuple[tuple[DataAssetRef | str, ...], tuple[str, ...]]:
        """Resolve and rehash caller-supplied refs before any output write."""

        normalized: list[DataAssetRef | str] = []
        hashes: list[str] = []
        for ref in source_refs:
            if isinstance(ref, DataAssetRef):
                source_path = self.context.resolve_input(ref.uri)
                if not source_path.is_file():
                    raise FileNotFoundError(source_path)
                actual_hash = _sha256_file(source_path)
                if ref.content_hash and ref.content_hash != actual_hash:
                    raise ValueError(f"source changed after registration: {ref.uri}")
                normalized.append(
                    DataAssetRef(
                        uri=str(source_path),
                        format=ref.format,
                        content_hash=actual_hash,
                        size_bytes=source_path.stat().st_size,
                        metadata=ref.metadata,
                    )
                )
                hashes.append(actual_hash)
                continue
            source_path = self.context.resolve_input(ref)
            if not source_path.is_file():
                raise FileNotFoundError(source_path)
            actual_hash = _sha256_file(source_path)
            normalized.append(str(source_path))
            hashes.append(actual_hash)
        return tuple(normalized), tuple(hashes)

    def _canonical_lineage(
        self,
        lineage: Mapping[str, Any] | None,
        *,
        member_lineage: Sequence[Mapping[str, Any]],
        transformations: Sequence[str],
    ) -> dict[str, Any]:
        provided = dict(lineage or {})
        reserved = sorted(_RESERVED_LINEAGE_KEYS.intersection(provided))
        if reserved:
            raise ValueError(f"lineage contains reserved keys: {', '.join(reserved)}")
        return {
            **provided,
            "archive": self._room.archive_ref.to_dict(),
            "members": list(member_lineage),
            "transformations": list(transformations),
        }

    def _materialize_rows(
        self,
        rows: Iterable[Mapping[str, Any]] | DataRoomMember | DataRoomCatalogEntry | str | Path,
        *,
        sheet: str | None,
        max_rows: int | None,
        limitations: list[str] | None = None,
    ) -> tuple[list[dict[str, Any]], list[DataRoomMember], str | None]:
        if max_rows is not None and (isinstance(max_rows, bool) or max_rows < 0):
            raise ValueError("max_rows cannot be negative")
        inferred_members: list[DataRoomMember] = []
        if isinstance(rows, (DataRoomMember, DataRoomCatalogEntry, str, Path)):
            catalog_entry = rows if isinstance(rows, DataRoomCatalogEntry) else None
            inferred_member = self._resolve_source_member(rows)
            inferred_members.append(inferred_member)
            if sheet is None and catalog_entry is not None:
                sheet = catalog_entry.sheet_name
            rows_value: Iterable[Mapping[str, Any]] = self._room.read_rows(
                # Keep the logical catalog view when a caller selected one
                # SQLite table.  Resolving only the physical member would
                # otherwise make read_rows default to the first table.
                rows if catalog_entry is not None else inferred_member,
                sheet=sheet,
                limit=None if max_rows is None else max_rows + 1,
            )
        else:
            rows_value = rows
        materialized: list[dict[str, Any]] = []
        truncated = False
        for row in rows_value:
            if not isinstance(row, Mapping):
                raise TypeError("prepared rows must be mappings")
            if max_rows is not None and len(materialized) >= max_rows:
                truncated = True
                break
            materialized.append(dict(row))
        if truncated and limitations is not None:
            limitations.append(f"prepared asset rows limited to explicit max_rows={max_rows}")
        return materialized, inferred_members, sheet

    @staticmethod
    def _infer_schema(rows: Sequence[Mapping[str, Any]], schema: Mapping[str, str] | None) -> dict[str, str]:
        if schema:
            return dict(schema)
        field_names: list[str] = []
        for row in rows:
            for key in row:
                if str(key) not in field_names:
                    field_names.append(str(key))
        return {
            key: next((_physical_type(row.get(key)) for row in rows if row.get(key) is not None), "null")
            for key in field_names
        }

    @staticmethod
    def _encode_prepared_rows(rows: Sequence[Mapping[str, Any]], normalized_format: str, schema: Mapping[str, str]) -> bytes:
        if normalized_format == "jsonl":
            return "".join(
                json.dumps(_jsonable(row), ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str) + "\n"
                for row in rows
            ).encode("utf-8")
        field_names = [str(key) for key in schema]
        for row in rows:
            for key in row:
                if str(key) not in field_names:
                    field_names.append(str(key))
        output = io.StringIO(newline="")
        writer = csv.DictWriter(output, fieldnames=field_names, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows({key: _cell_value(row.get(key)) for key in field_names} for row in rows)
        return output.getvalue().encode("utf-8")

    def _prepared_sources(
        self,
        inferred_members: Sequence[DataRoomMember],
        source_members: Iterable[DataRoomMember | DataRoomCatalogEntry | str | Path],
        source_refs: Iterable[DataAssetRef | str],
    ) -> tuple[tuple[DataRoomMember, ...], tuple[DataAssetRef | str, ...], tuple[str, ...]]:
        members = list(inferred_members)
        members.extend(self._resolve_source_member(value) for value in source_members)
        unique_members = tuple(dict((member.path, member) for member in members).values())
        normalized_refs, supplied_ref_hashes = self._normalize_source_refs(source_refs)
        source_hashes = [self._room.archive_ref.content_hash or ""]
        source_hashes.extend(member.content_hash for member in unique_members)
        source_hashes.extend(supplied_ref_hashes)
        source_hashes = tuple(dict.fromkeys(value for value in source_hashes if value))
        source_ref_values: tuple[DataAssetRef | str, ...] = (self._room.archive_ref, *normalized_refs)
        return unique_members, source_ref_values, source_hashes

    def _prepared_descriptor(
        self,
        *,
        prepared_asset_id: str,
        normalized_format: str,
        destination: Path,
        encoded: bytes,
        rows: Sequence[Mapping[str, Any]],
        schema: Mapping[str, str],
        grain: str | None,
        transformations: tuple[str, ...],
        identity_mappings: tuple[str, ...],
        relationship_mappings: tuple[str, ...],
        ontology_refs: tuple[str, ...],
        limitations: tuple[str, ...],
        lineage: Mapping[str, Any] | None,
        scope: str,
        unique_members: Sequence[DataRoomMember],
        source_refs: tuple[DataAssetRef | str, ...],
        source_hashes: tuple[str, ...],
        sheet: str | None,
        max_bytes: int | None,
        effective_period: str | None = None,
    ) -> PreparedAssetDescriptor:
        member_lineage = [
            {"path": member.path, "content_hash": member.content_hash, "format": member.format, "sheet": sheet}
            for member in unique_members
        ]
        lineage_payload = self._canonical_lineage(
            lineage,
            member_lineage=member_lineage,
            transformations=transformations,
        )
        manifest_payload = {
            "prepared_asset_id": prepared_asset_id,
            "format": normalized_format,
            "schema": dict(schema),
            "source_hashes": list(source_hashes),
            "transformations": list(transformations),
            "ontology_refs": list(ontology_refs),
            "effective_period": effective_period,
        }
        operation_manifest_hash = _sha256_bytes(json.dumps(manifest_payload, sort_keys=True, separators=(",", ":")).encode())
        return PreparedAssetDescriptor(
            prepared_asset_id=prepared_asset_id,
            source_refs=source_refs,
            source_hashes=source_hashes,
            location=str(destination),
            schema=dict(schema),
            grain=grain,
            transformations=transformations,
            identity_mappings=identity_mappings,
            relationship_mappings=relationship_mappings,
            ontology_refs=ontology_refs,
            limitations=limitations,
            lineage=lineage_payload,
            scope=scope,
            prepared_content_hash=_sha256_bytes(encoded),
            operation_manifest_hash=operation_manifest_hash,
            core_version=self.context.core_version,
            row_count=len(rows),
            byte_count=len(encoded),
            effective_period=effective_period,
            metadata={
                "format": normalized_format,
                "source_member_paths": [member.path for member in unique_members],
                "max_bytes": max_bytes,
            },
        )

    def _save_prepared_candidate(
        self,
        prepared_asset_id: str,
        rows: Iterable[Mapping[str, Any]] | DataRoomMember | DataRoomCatalogEntry | str | Path,
        *,
        candidate_root: Path,
        format: str = "jsonl",
        sheet: str | None = None,
        source_members: Iterable[DataRoomMember | DataRoomCatalogEntry | str | Path] = (),
        source_refs: Iterable[DataAssetRef | str] = (),
        schema: Mapping[str, str] | None = None,
        grain: str | None = None,
        transformations: Iterable[str] = (),
        identity_mappings: Iterable[str] = (),
        relationship_mappings: Iterable[str] = (),
        ontology_refs: Iterable[str] = (),
        limitations: Iterable[str] = (),
        lineage: Mapping[str, Any] | None = None,
        scope: str = "requirement_scoped",
        max_rows: int | None = None,
        max_bytes: int | None = DEFAULT_PREPARED_MAX_BYTES,
        max_output_bytes: int | None = None,
        effective_period: str | None = None,
    ) -> PreparedAssetDescriptor:
        prepared_asset_id = _safe_id(prepared_asset_id, "prepared_asset_id")
        normalized_format = str(format).lower().lstrip(".")
        if normalized_format == "ndjson":
            normalized_format = "jsonl"
        if normalized_format not in {"jsonl", "csv"}:
            raise ValueError("prepared assets support only JSONL or CSV")
        if scope not in PreparedAssetRegistry.allowed_scopes:
            raise ValueError(f"prepared asset scope is invalid: {scope!r}")
        if max_rows is not None and (isinstance(max_rows, bool) or max_rows < 0):
            raise ValueError("max_rows cannot be negative")
        if max_output_bytes is not None:
            if max_output_bytes < 0:
                raise ValueError("max_output_bytes cannot be negative")
            max_bytes = max_output_bytes
        if max_bytes is not None and (isinstance(max_bytes, bool) or max_bytes < 0):
            raise ValueError("max_bytes cannot be negative")
        self._room._check_archive_stat()
        source_members = tuple(source_members)
        source_refs = tuple(source_refs)
        transformations = tuple(str(value) for value in transformations)
        identity_mappings = tuple(str(value) for value in identity_mappings)
        relationship_mappings = tuple(str(value) for value in relationship_mappings)
        ontology_refs = tuple(str(value) for value in ontology_refs)
        limitations = tuple(str(value) for value in limitations)
        normalized_effective_period = None if effective_period is None else str(effective_period).strip()
        materialization_limitations: list[str] = []
        materialized, inferred_members, sheet = self._materialize_rows(
            rows,
            sheet=sheet,
            max_rows=max_rows,
            limitations=materialization_limitations,
        )
        limitations = tuple(dict.fromkeys((*limitations, *materialization_limitations)))
        unique_members, source_ref_values, source_hashes = self._prepared_sources(
            inferred_members,
            source_members,
            source_refs,
        )
        inferred_schema = self._infer_schema(materialized, schema)
        candidate_root = self.context.resolve_run_path(candidate_root)
        target_root = candidate_root / prepared_asset_id
        if not target_root.is_relative_to(candidate_root):
            raise ValueError("prepared candidate path escapes item workspace")
        filename = f"{prepared_asset_id}.{normalized_format}"
        destination = target_root / filename
        encoded = self._encode_prepared_rows(materialized, normalized_format, inferred_schema)
        if max_bytes is not None and len(encoded) > max_bytes:
            raise ValueError(f"prepared asset exceeds max_bytes: {len(encoded)} > {max_bytes}")
        descriptor = self._prepared_descriptor(
            prepared_asset_id=prepared_asset_id,
            normalized_format=normalized_format,
            destination=destination,
            encoded=encoded,
            rows=materialized,
            schema=inferred_schema,
            grain=grain,
            transformations=transformations,
            identity_mappings=identity_mappings,
            relationship_mappings=relationship_mappings,
            ontology_refs=ontology_refs,
            limitations=limitations,
            lineage=lineage,
            scope=scope,
            unique_members=unique_members,
            source_refs=source_ref_values,
            source_hashes=source_hashes,
            sheet=sheet,
            max_bytes=max_bytes,
            effective_period=normalized_effective_period,
        )
        descriptor_path = target_root / f"{prepared_asset_id}.descriptor.json"
        sidecar_bytes = (json.dumps(descriptor.to_dict(), ensure_ascii=False, sort_keys=True, indent=2) + "\n").encode("utf-8")
        lock_path = candidate_root / ".candidates.lock"
        with _candidate_lock(lock_path):
            if target_root.exists():
                if target_root.is_symlink() or not target_root.is_dir():
                    raise ValueError(f"prepared candidate residue is not a directory: {target_root}")
                if not destination.is_file() or destination.is_symlink() or not descriptor_path.is_file() or descriptor_path.is_symlink():
                    raise ValueError(f"prepared candidate residue is incomplete: {prepared_asset_id}")
                try:
                    existing = PreparedAssetDescriptor.from_dict(json.loads(descriptor_path.read_text(encoding="utf-8")))
                except (OSError, UnicodeDecodeError, json.JSONDecodeError, TypeError, ValueError, KeyError) as exc:
                    raise ValueError(f"prepared candidate descriptor is invalid: {prepared_asset_id}") from exc
                if existing != descriptor or destination.read_bytes() != encoded:
                    raise ValueError(f"prepared candidate already exists with different descriptor: {prepared_asset_id}")
                descriptor = existing
            else:
                residue = tuple(candidate_root.glob(f".{prepared_asset_id}.staging-*"))
                if residue:
                    raise ValueError(f"prepared candidate staging residue exists: {prepared_asset_id}")
                temporary_root = candidate_root / f".{prepared_asset_id}.staging-{uuid.uuid4().hex}"
                temporary_root.mkdir(parents=True, exist_ok=False)
                try:
                    temporary_destination = temporary_root / filename
                    temporary_descriptor = temporary_root / f"{prepared_asset_id}.descriptor.json"
                    _atomic_write_bytes(temporary_destination, encoded)
                    _atomic_write_bytes(temporary_descriptor, sidecar_bytes)
                    directory = os.open(temporary_root, os.O_RDONLY)
                    try:
                        try:
                            os.fsync(directory)
                        except OSError:
                            pass
                    finally:
                        os.close(directory)
                    os.replace(temporary_root, target_root)
                    try:
                        parent_fd = os.open(candidate_root, os.O_RDONLY)
                    except OSError:
                        parent_fd = None
                    if parent_fd is not None:
                        try:
                            try:
                                os.fsync(parent_fd)
                            except OSError:
                                pass
                        finally:
                            os.close(parent_fd)
                finally:
                    if temporary_root.exists():
                        shutil.rmtree(temporary_root, ignore_errors=True)
        self._prepared[prepared_asset_id] = descriptor
        self._emit_prepared(descriptor)
        return descriptor

    def _emit_prepared(self, descriptor: PreparedAssetDescriptor) -> None:
        try:
            self.telemetry.record(
                "data_room_prepared_write",
                bytes_processed=descriptor.byte_count,
                rows=descriptor.row_count,
                output_hashes=(descriptor.prepared_content_hash,) if descriptor.prepared_content_hash else (),
                facts={"prepared_asset_id": descriptor.prepared_asset_id, "location": descriptor.location},
            )
        except Exception:
            pass

    def prepared(self, prepared_asset_id: str) -> PreparedAsset:
        prepared_asset_id = _safe_id(prepared_asset_id, "prepared_asset_id")
        loaded = self.prepared_registry.load(prepared_asset_id)
        self._prepared[prepared_asset_id] = loaded.descriptor
        return loaded

__all__ = [
    "CatalogCounts",
    "DataRoom",
    "DataRoomCatalogEntry",
    "DataRoomMember",
    "DataRoomWorkbench",
    "PreparedAsset",
    "PreparedAssetRegistry",
]
