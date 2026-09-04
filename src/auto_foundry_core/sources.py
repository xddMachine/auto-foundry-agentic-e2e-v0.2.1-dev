"""Read-only local source registration, discovery, previews, and bounded reads."""

from __future__ import annotations

import csv
from dataclasses import asdict
import hashlib
import json
from pathlib import Path
import sqlite3
from typing import Any, Iterable, Iterator, Mapping
from urllib.parse import quote

from .contracts import DataAssetRef, DocumentRef, TableRef
from .workspace import validate_allowed_path


TABULAR_FORMATS = frozenset({"csv", "tsv", "json", "jsonl", "ndjson", "xlsx", "parquet", "sqlite"})
SQLITE_SUFFIXES = frozenset({"db", "sqlite", "sqlite3"})
TEXT_FORMATS = frozenset({"txt", "md", "markdown", "rst", "html", "htm", "xml", "log"})
# JSON is decoded without an arbitrary default byte ceiling.  Callers that
# need a bounded operation may provide ``max_json_bytes`` explicitly.
DEFAULT_JSON_MAX_BYTES: int | None = None


def hash_file(
    path: str | Path,
    *,
    chunk_size: int = 1024 * 1024,
    allowed_roots: Iterable[str | Path] | None = None,
) -> str:
    """Return a stable SHA-256 hash without loading the whole file."""

    path = _bounded(path, allowed_roots)
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        while chunk := stream.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _format(path: Path, explicit: str | None = None) -> str:
    value = explicit or path.suffix.lstrip(".") or "binary"
    value = value.lower().lstrip(".")
    if value == "ndjson":
        return "jsonl"
    if value in SQLITE_SUFFIXES:
        return "sqlite"
    return value


def _bounded(path: str | Path, allowed_roots: Iterable[str | Path] | None) -> Path:
    candidate = Path(path).expanduser().resolve(strict=False)
    if allowed_roots is not None:
        candidate = validate_allowed_path(candidate, allowed_roots)
    return candidate


def register_source(
    path: str | Path,
    *,
    allowed_roots: Iterable[str | Path] | None = None,
    format: str | None = None,
    metadata: Mapping[str, Any] | None = None,
) -> DataAssetRef:
    """Register a local source without changing it."""

    candidate = _bounded(path, allowed_roots)
    if not candidate.exists() or not candidate.is_file():
        raise FileNotFoundError(candidate)
    return DataAssetRef(
        uri=str(candidate),
        format=_format(candidate, format),
        content_hash=hash_file(candidate, allowed_roots=allowed_roots),
        size_bytes=candidate.stat().st_size,
        metadata={"read_only": True, **dict(metadata or {})},
    )


def register_document(
    path: str | Path,
    *,
    allowed_roots: Iterable[str | Path] | None = None,
    mime_type: str | None = None,
    title: str | None = None,
) -> DocumentRef:
    asset = register_source(path, allowed_roots=allowed_roots)
    return DocumentRef(asset=asset, title=title or Path(asset.uri).name, mime_type=mime_type)


def _asset(source: DataAssetRef | str | Path, allowed_roots: Iterable[str | Path] | None = None) -> DataAssetRef:
    if isinstance(source, DataAssetRef):
        path = _bounded(source.uri, allowed_roots)
        if source.content_hash and path.is_file() and hash_file(path, allowed_roots=allowed_roots) != source.content_hash:
            raise ValueError(f"source changed after registration: {path}")
        # Re-hash only when caller passed a descriptor that does not carry one.
        if source.content_hash:
            return source
        return register_source(path, allowed_roots=allowed_roots, format=source.format, metadata=source.metadata)
    return register_source(source, allowed_roots=allowed_roots)


def _openpyxl():
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("Excel support requires the optional 'openpyxl' dependency") from exc
    return openpyxl


def _pyarrow():
    try:
        import pyarrow.parquet as pq
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("Parquet support requires the optional 'pyarrow' dependency") from exc
    return pq


def _quote_sqlite_identifier(identifier: str) -> str:
    """Quote one SQLite identifier without allowing SQL fragments."""

    value = str(identifier)
    if not value or "\x00" in value:
        raise ValueError("SQLite identifiers must be non-empty and NUL-free")
    return '"' + value.replace('"', '""') + '"'


def _sqlite_connect(path: Path) -> sqlite3.Connection:
    """Open a SQLite database read-only and force query-only mode."""

    # URI mode=ro prevents accidental creation or writes.  Quote the complete
    # path so ``?``, ``#`` and other URI-significant bytes cannot alter the
    # connection flags.
    uri = f"file:{quote(str(path), safe='/:')}?mode=ro"
    connection = sqlite3.connect(uri, uri=True)
    try:
        connection.execute("PRAGMA query_only = ON")
        connection.row_factory = sqlite3.Row
    except Exception:
        connection.close()
        raise
    return connection


def _sqlite_tables(path: Path) -> list[str]:
    """Return user tables in deterministic binary-name order."""

    connection = _sqlite_connect(path)
    try:
        rows = connection.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type = 'table' AND name NOT LIKE 'sqlite_%' "
            "ORDER BY name COLLATE BINARY"
        ).fetchall()
        return [str(row[0]) for row in rows]
    finally:
        connection.close()


def _sqlite_table_name(path: Path, table: str | None) -> str:
    names = _sqlite_tables(path)
    if table is None:
        if not names:
            raise ValueError(f"SQLite database contains no user tables: {path.name}")
        return names[0]
    value = str(table)
    if value not in names:
        raise KeyError(f"unknown SQLite table: {value}")
    return value


def discover(
    source: DataAssetRef | str | Path,
    *,
    allowed_roots: Iterable[str | Path] | None = None,
) -> list[TableRef | DocumentRef]:
    """Discover logical tables/sheets without reading unbounded data."""

    asset = _asset(source, allowed_roots)
    path = Path(asset.uri)
    fmt = _format(path, asset.format)
    if fmt == "xls":
        raise ValueError("unsupported source format: xls; use xlsx or convert the workbook explicitly")
    if fmt == "xlsx":
        workbook = _openpyxl().load_workbook(path, read_only=True, data_only=True)
        try:
            return [TableRef(asset=asset, name=str(name), kind="sheet") for name in workbook.sheetnames]
        finally:
            workbook.close()
    if fmt == "sqlite":
        return [TableRef(asset=asset, name=name, kind="table") for name in _sqlite_tables(path)]
    if fmt in TABULAR_FORMATS:
        return [TableRef(asset=asset, name=path.stem, kind="table")]
    return [DocumentRef(asset=asset, title=path.name)]


def _rows_csv(path: Path, *, delimiter: str = ",") -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        for row in csv.DictReader(stream, delimiter=delimiter):
            yield dict(row)


def _rows_json(path: Path, fmt: str, *, max_bytes: int | None = DEFAULT_JSON_MAX_BYTES) -> Iterator[dict[str, Any]]:
    if fmt == "jsonl":
        with path.open("r", encoding="utf-8-sig") as stream:
            for line_number, line in enumerate(stream, 1):
                if not line.strip():
                    continue
                value = json.loads(line)
                yield value if isinstance(value, dict) else {"value": value, "_line": line_number}
        return
    if max_bytes is not None and (isinstance(max_bytes, bool) or max_bytes < 0):
        raise ValueError("max_json_bytes cannot be negative")
    size_bytes = path.stat().st_size
    if max_bytes is not None and size_bytes > max_bytes:
        raise ValueError(
            f"ordinary JSON materialization boundary exceeded: {size_bytes} bytes > {max_bytes} max_json_bytes"
        )
    with path.open("r", encoding="utf-8-sig") as stream:
        value = json.load(stream)
    if isinstance(value, dict):
        # A conventional records container is convenient, while preserving a
        # scalar/object document as a single row.
        records = value.get("records", value.get("data"))
        if isinstance(records, list):
            value = records
        else:
            value = [value]
    if not isinstance(value, list):
        value = [value]
    for item in value:
        yield item if isinstance(item, dict) else {"value": item}


def _rows_excel(path: Path, sheet: str | None = None) -> Iterator[dict[str, Any]]:
    workbook = _openpyxl().load_workbook(path, read_only=True, data_only=True)
    try:
        name = sheet or workbook.sheetnames[0]
        if name not in workbook.sheetnames:
            raise KeyError(f"unknown worksheet: {name}")
        values = workbook[name].iter_rows(values_only=True)
        try:
            headers = [str(v) if v is not None else f"column_{i + 1}" for i, v in enumerate(next(values))]
        except StopIteration:
            return
        for row in values:
            yield {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
    finally:
        workbook.close()


def _rows_parquet(path: Path, *, batch_size: int = 1024) -> Iterator[dict[str, Any]]:
    if batch_size <= 0:
        raise ValueError("parquet_batch_size must be positive")
    parquet = _pyarrow().ParquetFile(path)
    # iter_batches reads one bounded batch at a time.  In particular, this
    # avoids pyarrow.read_table(), which materializes every row group before
    # read_rows can apply offset/limit.
    for batch in parquet.iter_batches(batch_size=batch_size):
        yield from batch.to_pylist()


def _rows_sqlite(path: Path, table: str | None, *, batch_size: int = 1024) -> Iterator[dict[str, Any]]:
    if batch_size <= 0:
        raise ValueError("sqlite_batch_size must be positive")
    name = _sqlite_table_name(path, table)
    connection = _sqlite_connect(path)
    try:
        cursor = connection.execute(f"SELECT * FROM {_quote_sqlite_identifier(name)}")
        while True:
            batch = cursor.fetchmany(batch_size)
            if not batch:
                break
            for row in batch:
                yield {str(key): row[key] for key in row.keys()}
    finally:
        connection.close()


def _rows_text(path: Path) -> Iterator[dict[str, Any]]:
    with path.open("r", encoding="utf-8", errors="replace") as stream:
        for line in stream:
            yield {"text": line.rstrip("\n")}


def read_rows(
    source: DataAssetRef | TableRef | str | Path,
    *,
    table: str | None = None,
    limit: int | None = None,
    offset: int = 0,
    max_json_bytes: int | None = DEFAULT_JSON_MAX_BYTES,
    parquet_batch_size: int = 1024,
    allowed_roots: Iterable[str | Path] | None = None,
) -> list[dict[str, Any]]:
    """Read at most ``limit`` records, materializing only the bounded slice."""

    if limit is not None and limit < 0:
        raise ValueError("limit cannot be negative")
    if offset < 0:
        raise ValueError("offset cannot be negative")
    if isinstance(source, TableRef):
        sheet = table or source.name
        source = source.asset
        table = sheet
    asset = _asset(source, allowed_roots)
    path = Path(asset.uri)
    fmt = _format(path, asset.format)
    if fmt == "csv":
        rows = _rows_csv(path)
    elif fmt == "tsv":
        rows = _rows_csv(path, delimiter="\t")
    elif fmt in {"json", "jsonl"}:
        rows = _rows_json(path, fmt, max_bytes=max_json_bytes)
    elif fmt == "xlsx":
        rows = _rows_excel(path, table)
    elif fmt == "xls":
        raise ValueError("unsupported source format: xls; use xlsx or convert the workbook explicitly")
    elif fmt == "parquet":
        rows = _rows_parquet(path, batch_size=parquet_batch_size)
    elif fmt == "sqlite":
        rows = _rows_sqlite(path, table)
    elif fmt in TEXT_FORMATS:
        rows = _rows_text(path)
    else:
        raise ValueError(f"unsupported source format: {fmt}")
    output: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        if index < offset:
            continue
        output.append(dict(row))
        if limit is not None and len(output) >= limit:
            break
    return output


def preview(
    source: DataAssetRef | TableRef | str | Path,
    *,
    limit: int = 20,
    max_json_bytes: int | None = DEFAULT_JSON_MAX_BYTES,
    parquet_batch_size: int = 1024,
    allowed_roots: Iterable[str | Path] | None = None,
) -> dict[str, Any]:
    """Return a bounded, JSON-serializable source preview."""

    if limit < 0:
        raise ValueError("limit cannot be negative")
    asset = source.asset if isinstance(source, TableRef) else source
    if isinstance(asset, DataAssetRef):
        ref = _asset(asset, allowed_roots)
    else:
        ref = _asset(asset, allowed_roots)
    rows = read_rows(
        source,
        limit=limit,
        max_json_bytes=max_json_bytes,
        parquet_batch_size=parquet_batch_size,
        allowed_roots=allowed_roots,
    )
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return {
        "source": ref.to_dict(),
        "format": ref.format,
        "columns": columns,
        "sample": rows,
        "bounded": True,
    }


register = register_source
read = read_rows
source_hash = hash_file

__all__ = [
    "DEFAULT_JSON_MAX_BYTES",
    "TABULAR_FORMATS",
    "discover",
    "hash_file",
    "preview",
    "read",
    "read_rows",
    "register",
    "register_document",
    "register_source",
    "source_hash",
]
